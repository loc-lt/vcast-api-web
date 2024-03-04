from flask import *
import pandas as pd
import numpy as np
import math
import pyodbc
import io
import datetime
import requests
import openpyxl
import shutil
import traceback
from openpyxl import load_workbook,styles
from openpyxl.styles import Alignment
from openpyxl.styles import Color, Fill, Border, Side
from openpyxl.styles import Font
import warnings
import matplotlib.pyplot as plt
from flask_jwt_extended import JWTManager
from fastapi import FastAPI
import os
from os import listdir
import matplotlib.dates as mdates
from flask import session
import cv2
def configdata(data):
    for i in range(1,len(data)-1):
        while len(data[i]) < len(data[0]):
            data[i].append("")
    return data[:-1]
def calculate_standard_deviation(K107):
    sqrt_term = math.sqrt(2 / (K107 - 1))
    gamma_ln1 = math.lgamma(K107 / 2)
    gamma_ln2 = math.lgamma((K107 - 1) / 2)
    result = math.exp(math.log(sqrt_term) + gamma_ln1 - gamma_ln2)
    return result
def check_secr(tg):
    scr = str(request.cookies.get("scr"))
    if scr == "1" or bin(int(scr))[::-1][tg] == "1":
        return True
    else:
        return False
app = Flask(__name__, instance_relative_config=True)
app.config['SECRET_KEY'] = 'thisissecret'
JWTManager(app)
@app.route("/login", methods=['POST', 'GET'])
def login():
    if request.method == 'POST':        
        username = request.form['username']
        password = request.form['password']
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=Auto; Trusted_Connection=No;',timeout=1)
            # data = laser_all_data("","","",timestart,timeend)
        get_user = pd.read_sql("SELECT Top(1) * FROM [Auto].[dbo].[Web_User_Test] where username = '"+username+"'", conn)
        if len(get_user) == 1:
            confirm_user = pd.read_sql("SELECT Top(1) * FROM [Auto].[dbo].[Web_User_Test] where username = '"+username+"' and password = '"+password+"'" , conn)
            print(confirm_user)
            if len(confirm_user) == 1:
                scr = pd.read_sql("SELECT security FROM [Auto].[dbo].[Web_Role] where role = '"+confirm_user.loc[0]['role']+"'" , conn).loc[0,"security"]
                resp = make_response(redirect(url_for('hello_world')))
                resp.set_cookie('login_status', 'ok')   
                resp.set_cookie('user_name', confirm_user.loc[0]['username']) 
                resp.set_cookie('scr', str(scr)) 
                return resp
            else:
                return render_template('pages/login.html',stt = 'Wrong Password')
        else:
            return render_template('pages/login.html',stt = 'No User Found')
    else:
        resp = make_response(render_template('pages/login.html'))
        resp.set_cookie('login_status', '')
        resp.set_cookie('user_name', '')
        resp.set_cookie('scr', '') 
        return resp
@app.route("/product", methods=['GET'])
def product():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")   
        if check_secr(1):            
            product_list = listdir("static/product")
            return render_template('pages/product.html',user=user,process = "Product",product_list = product_list)
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/product/<product>", methods=['GET'])
def product_detail(product):
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get( "user_name") 
        if check_secr(1):    
            if product == "A2012003" or product == "A2012004":      
                lst_banve = []
                count = 0
                for x in os.listdir("static/product/"+product+"/02. DRAWING"):
                    if "pdf" in x :
                        lst_banve.append({"name":x,"number":count})
                        count+=1
                lst_sodo = []
                count = 0
                for x in os.listdir("static/product/"+product+"/03. PROCESS FLOW CHART"):
                    if "pdf" in x :
                        lst_sodo.append({"name":x,"number":count})
                        count+=1
                return render_template('pages/product_detail.html',lst_banve=lst_banve,lst_sodo=lst_sodo,user=user,process = "Product",product = product)
            else:
                return render_template('pages/feature.html',user=user,process="Product",product=product)
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route('/product/<product>/<filestyle>/<number>')
def download_pdf(product,filestyle,number):
    # Đường dẫn tới file PDF
    if filestyle == "drawing":    
        print("hello")
        # file_path = 'path/to/your/pdf/file.pdf'
        folder_path = f'static/product/{product}/02. DRAWING/'
    elif filestyle == "flowchart":
        folder_path = f'static/product/{product}/03. PROCESS FLOW CHART/'
    pdf_files = listdir(folder_path)
    lst_pdf = []
    for i in pdf_files:
        if "pdf" in i:
            lst_pdf.append(i)
    return send_file(folder_path+lst_pdf[int(number)])
@app.route("/", methods=['POST', 'GET'])
def hello_world():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if request.method == 'POST':
            DMC=request.form['DMC']
            conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
            # data = laser_all_data("","","",timestart,timeend)
            cursor = conn.cursor()
            cursor.execute("SELECT * From[QC].[dbo].[Laser] where DMCout = '"+DMC+"'")
            laser_data =cursor.fetchall()
            cursor.execute("SELECT * From[QC].[dbo].[CNC] where DMC_product like '%"+DMC+"%'")
            CNC_data =cursor.fetchall()
            cursor.execute("SELECT * From[QC].[dbo].[air_tight] where Barcode= '"+DMC+"'")
            Airtight_data =cursor.fetchall()
            cursor.execute("SELECT * From[QC].[dbo].[Classification] where total_code = '"+DMC+"'")
            Classification_data =cursor.fetchall()
            # print(data['DMCin'])
            return render_template('pages/component.html',user=user,DMC=DMC,laser_data=laser_data,CNC_data=CNC_data,Airtight_data=Airtight_data,Classification_data=Classification_data)
        else:
            return render_template('pages/show.html',user=user,process = "Home")
    else:
        return redirect(url_for('login'))
@app.route("/Search", methods=['POST', 'GET'])
def Search():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(11): 
            if request.method == 'POST':
                DMC=request.form['DMC']
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                # data = laser_all_data("","","",timestart,timeend)
                cursor = conn.cursor()
                if len(DMC)==29:
                    ctc = DMC[21:28]
                    if ctc[1] == 'X':
                        mathang = '10'
                    elif ctc[1] == 'Y':
                        mathang = '11'
                    elif ctc[1] == 'Z':
                        mathang = '12'
                    else:
                        mathang = '0'+ctc[1]
                    Castingcode = '202'+ctc[0]+'.'+mathang+'.'+ctc[2:4]+'-'+ctc[4:]
                else:
                    Castingcode = ''
                # cursor.execute("SELECT * From [SPC].[dbo].spc6() where CastingNo = '"+Castingcode+"'")
                # print("SELECT top 1 * From [SPC].[dbo].spc6() where CastingNo = '"+Castingcode+"'")
                # opticaldivision_data =cursor.fetchall()

                cursor.execute("SELECT * From[QC].[dbo].[Laser] where DMCout = '"+DMC+"'")
                laser_data =cursor.fetchall()

                cursor.execute("SELECT * From[QC].[dbo].[ScanQR] where DMC_product = '"+DMC+"'")
                ScanQR_data =cursor.fetchall()

                cursor.execute("SELECT * From[QC].[dbo].[TiltMeasurement] where DMC = '"+DMC+"'")
                TiltMeasurement_data =cursor.fetchall()

                cursor.execute("SELECT * From[QC].[dbo].[CNC] where DMC_product like '%"+DMC+"%'")
                CNC_data =cursor.fetchall()

                cursor.execute("select * from  QC.dbo.air_tight  WHERE barcode = '"+DMC+"'")
                Airtight_data =cursor.fetchall()
                
                cursor.execute("select * from  QC.dbo.air_tight_chamfer  WHERE barcode = '"+DMC+"'")
                Airtight_chamfer_data =cursor.fetchall()

                cursor.execute("select * from  QC.dbo.air_tight_window  WHERE barcode = '"+DMC+"'")
                Airtight_window_data =cursor.fetchall()
                # cursor.execute("SELECT * From[QC].[dbo].[air_tight] where Barcode= '"+DMC+"'")
                # Airtight_data =cursor.fetchall()

                # cursor.execute("SELECT * From[QC_SWVN].[dbo].[airtight_StrongWay] where Barcode= '"+DMC+"'")
                # AirtightSW_data =cursor.fetchall()

                cursor.execute("SELECT * From[QC].[dbo].[ThreadVerification] where DMC_product = '"+DMC+"'")
                ThreadVerification_data =cursor.fetchall()

                

                cursor.execute("SELECT * From[QC].[dbo].[Measure_Diameter] where DMC = '"+DMC+"'")
                AirguageQC_data =cursor.fetchall()

                cursor.execute("SELECT * From[GC].[dbo].[Measure_Diameter] where DMC = '"+DMC+"'")
                AirguageGC_data =cursor.fetchall()

                cursor.execute("SELECT [Pallet].[SPLR_LOT_NO], [Classification].* From [Classification] left join [Pallet] on [Classification].pallet_num= [Pallet].Pallet_Name where total_code = '"+DMC+"'")
                Classification_data =cursor.fetchall()
                # print(data['DMCin'])
                return render_template('pages/Search.html',user=user,process = "Search",post_flag = True,DMC=DMC,Airtight_window_data=Airtight_window_data,Airtight_chamfer_data=Airtight_chamfer_data,laser_data=laser_data,ScanQR_data=ScanQR_data,TiltMeasurement_data=TiltMeasurement_data,CNC_data=CNC_data,ThreadVerification_data=ThreadVerification_data,AirguageGC_data=AirguageGC_data,AirguageQC_data=AirguageQC_data, Airtight_data=Airtight_data,Classification_data=Classification_data)
            else:
                return render_template('pages/Search.html',user=user,process = "Search")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))

@app.route("/CNCdelete", methods=['POST', 'GET'])
def CNCdelete():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(26): 
            if request.method == 'POST':
                DMC=request.form['DMC']
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                # data = laser_all_data("","","",timestart,timeend)
                cursor = conn.cursor()
                if len(DMC)==29:
                    ctc = DMC[21:28]
                    if ctc[1] == 'X':
                        mathang = '10'
                    elif ctc[1] == 'Y':
                        mathang = '11'
                    elif ctc[1] == 'Z':
                        mathang = '12'
                    else:
                        mathang = '0'+ctc[1]
                    Castingcode = '202'+ctc[0]+'.'+mathang+'.'+ctc[2:4]+'-'+ctc[4:]
                else:
                    Castingcode = ''
       
                cursor.execute("SELECT * From[QC].[dbo].[ScanQR] where DMC_product = '"+DMC+"'")
                ScanQR_data =cursor.fetchall()

                cursor.execute("SELECT * From[QC].[dbo].[TiltMeasurement] where DMC = '"+DMC+"'")
                TiltMeasurement_data =cursor.fetchall()

                button = 0
                if len(ScanQR_data) >0 or len(TiltMeasurement_data)>0:
                    button =1

                return render_template('pages/CNCdelete.html',user=user,process = "CNCdelete",post_flag = True,button=button,DMC=DMC,ScanQR_data=ScanQR_data,TiltMeasurement_data=TiltMeasurement_data)
            else:
                return render_template('pages/CNCdelete.html',user=user,process = "CNCdelete")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/CNC_edit", methods=['POST'])
def CNC_edit():
    if check_secr(26): 
        username = request.cookies.get("user_name")
        data = request.get_json()
        dmc = data.get("dmc")
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        scandelete = pd.read_sql("SELECT  * FROM [QC].[dbo].[ScanQR] where DMC_product = '"+dmc+"'" , conn)
        tiltdelete = pd.read_sql("SELECT  * FROM [QC].[dbo].[TiltMeasurement] where DMC = '"+dmc+"'" , conn)
        time = str(datetime.datetime.now())[:-3]
        timedelete = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        cursor = conn.cursor()
        try:
            for index, row in scandelete.iterrows():
                insert_sql = "INSERT INTO [ScanQRdelete] ([User_delete],[Time_delete],[Product], [Time_scan_product], [Time_scan_tray], [DMC_product], [DMC_tray], [Compare]) VALUES (?, ?, ?, ?, ?, ?, ?, ?)"
                cursor.execute(insert_sql, (username,timedelete,row['Product'], row['Time_scan_product'], row['Time_scan_tray'], row['DMC_product'], row['DMC_tray'], row['Compare']))
            conn.commit()
            print("Dữ liệu scanqr đã được chèn thành công.")

            for index, row in tiltdelete.iterrows():
                insert_sql = "INSERT INTO [TiltMeasurement_delete] ([User_delete],[Time_delete],[ID_Operator],[MachineNo],[Nameproduct],[QR_tray],[DMC],[Time_start],[Time_finish],[Total_deviation_height_of_A_datum],[Distance_from_M_datum_to_line_of_2_pins_center],[Angle_of_part_vs_master],[Height_1_of_A_datum],[Height_2_of_A_datum],[Height_3_of_A_datum],[Height_4_of_A_datum],[Height_5_of_A_datum],[Result],[Status],[Picture],[Note]) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
                cursor.execute(insert_sql, (username,timedelete,row['ID_Operator'],row['MachineNo'],row['Nameproduct'],row['QR_tray'],row['DMC'],row['Time_start'],row['Time_finish'],row['Total_deviation_height_of_A_datum'],row['Distance_from_M_datum_to_line_of_2_pins_center'],row['Angle_of_part_vs_master'],row['Height_1_of_A_datum'],row['Height_2_of_A_datum'],row['Height_3_of_A_datum'],row['Height_4_of_A_datum'],row['Height_5_of_A_datum'],row['Result'],row['Status'],row['Picture'],row['Note']))
            conn.commit()
            print("Dữ liệu đo độ nghiêng đã được chèn thành công.")
        except Exception as e:
            print(f"Lỗi: {str(e)}")
        # conn.close()
        
        
        cursor.execute("delete from [QC].[dbo].[ScanQR] where [DMC_product] = '"+dmc+"' ")
        cursor.execute("delete from [QC].[dbo].[TiltMeasurement] where DMC = '"+dmc+"' ")
        # print("delete from [QC].[dbo].[ScanQR] where [DMC_product] = '"+dmc+"' ")
        # print("delete from [QC].[dbo].[TiltMeasurement] where DMC = '"+dmc+"' ")
        cursor.execute("insert into [QC].[dbo].[CNCdelete_history] ([ten],[thoigianxoa],[DMC]) values('"+username+"', '"+time+"','"+dmc+"')")
        cursor.commit()
        return 'OK'

    else:
        return "You don't have permission"
@app.route('/download/CNCx4',methods=['POST'])
def download_CNCx4():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT Top(15000)*  From[GC].[dbo].[ProductControl] where TimeScan >='"+timestart+"' and TimeScan <='"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=CNCx4_Data.xls"})
@app.route("/CNCx4", methods=['POST', 'GET'])
def open_CNCx4():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="CNC"
        if check_secr(17): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)*  From[GC].[dbo].[ProductControl] where TimeScan >='"+timestart+"' and TimeScan <='"+timeend+"'")
                
                # cursor = conn.cursor()
                # cursor.execute("SELECT Top(15000)*  From[QC].[dbo].[CNC] where TimeoutCNC1 >='"+timestart+"' and TimeoutCNC1 <='"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                # pagecount = int(soluong/100)
                # print(data['DMCin'])
                return render_template('pages/CNCx4.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/CNCx4.html',user=user,process=process,timestart="",timeend="")  
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/OpticalDivision", methods=['POST', 'GET'])
def Searchoptical():
    
    # global datafileid_str, rs2_list
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(4): 
            if request.method == 'POST':
                # DMC=request.form['DMC']   
                # Castingcode = request.form['Castingcode']
                # Productcode = request.form['Productcode']
                
                # error_flag = False
                timestart=request.form['daystart']+" "+request.form['timestart']
                if request.form['daystart'] == '':
                    timestart = '1900-01-01' + timestart
                timeend=request.form['dayend']+" "+request.form['timeend']
                if request.form['dayend'] == '':
                    timeend = '2100-01-01'+ timeend
                filegroup=request.form['grpname']
                filename=request.form['filename']
                print('aaaaaaa',filegroup)
                print('bbbbbbb',filename)
                

                conn1 = pyodbc.connect('Driver={SQL Server}; Server=192.168.9.246; uid=sa; pwd=saa;Database=SPC; Trusted_Connection=No;',timeout=1)
                # group = "select GroupName from FileGroup_V WHERE GroupName LIKE '%DC%' or GroupName Like '%DS%' or GroupName Like '%GC%' or GroupName Like '%Engineer%' or GroupName like '%VC-Test%' ORDER BY GroupName"
                group = "select GroupName from FileGroup_V WHERE GroupName ='DC10200-Casting' "
                cursor1 = conn1.cursor()   
                cursor1.execute(group)
                groupname= cursor1.fetchall()
                datafileid = session.get('datafileid')
                print('fileid0',datafileid)
                datafileid_str = str(datafileid["data"][0])
                print('fileid1',datafileid_str)
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=SPC; Trusted_Connection=No;',timeout=1)
                cursor = conn.cursor()            
                # cmd = "select * from SPC6() where FileID = '"+datafileid_str+"' and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'"
                
                if filegroup =='DS50000-Heat-Treatment':
                    cmdtorque = "select * from SPCtorque() where FileID = '"+datafileid_str+"' and VCTRLID in (SELECT [VCTRLID] FROM [SPC].[dbo].[HeatTreament_Info]) and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'"
                    cursor.execute(cmdtorque)
                    opticaldivision_data =cursor.fetchall()
                elif filegroup =='DC10200-Casting':  
                    cmd="SELECT * FROM dbo.newoptical('"+datafileid_str+"', '"+timestart+"', '"+timeend+"')"
                    # cmdtorque = "select * from SPCtorque() where FileID = '"+datafileid_str+"' and VCTRLID in (SELECT [VCTRLID] FROM [SPC].[dbo].[HeatTreament_Info]) and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'"
                    print(cmd)  
                    cursor.execute(cmd)
                    opticaldivision_data =cursor.fetchall()
                
                print('torque_data',opticaldivision_data)
                
                # print(opticaldivision_data)
                if filegroup =='DS50000-Heat-Treatment':
                    data1 = pd.read_sql("SELECT * FROM GetControltorque() where fileid = '"+datafileid_str+"'",conn1)
                    print()
                    data2 = pd.read_sql("select * from spc.dbo.getcontrol821_torque('"+datafileid_str+"', '"+timestart+"', '"+timeend+"')",conn)
                    data3 = pd.read_sql("SELECT * FROM spctorque() where fileid = '"+datafileid_str+"' and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'",conn)
                    post = 0
                elif filegroup =='DC10200-Casting':
                    data1 = pd.read_sql("SELECT * FROM GetControl3() where fileid = '"+datafileid_str+"'",conn1)
                    dff=pd.read_sql("SELECT * FROM dbo.newoptical('"+datafileid_str+"', '"+timestart+"', '"+timeend+"')",conn)
                    df = dff.drop('DateTimeCreate', axis=1)
                    df1 = df.agg(['max', 'min', 'mean'])
                    df2 = pd.DataFrame(df1, index=['max', 'min', 'mean'])
                    df2['Control_Item'] = df2.index
                    # df2[df2.select_dtypes(include='float').columns] = df2[df2.select_dtypes(include='float').columns].round(5)
                    data2 = df2[['Control_Item'] + [col for col in df2.columns if col != 'Control_Item']]
                    data3 = pd.read_sql("SELECT * FROM dbo.newoptical('"+datafileid_str+"', '"+timestart+"', '"+timeend+"')",conn)
                    post = 1
                
                common_columns = data1.columns.intersection(data2.columns)
                rs = pd.concat([data1[common_columns], data2[common_columns]])
                # rs = pd.concat([data1, data2], ignore_index=True)
                print('rs',rs)
                newdata = {}
                try:
                    for col in rs.columns[1:]:
                        newdata[col] = []
                        splqlt = len(data3) -1
                        test = 0
                        print('a1',rs.loc[4,col] )
                        sps = int(rs.loc[4,col])
                        if sps == 1:
                            try:
                                for i in range(splqlt):
                                    test += abs(data3.loc[i+1,col] - data3.loc[i,col])
                                dev = test/splqlt/1.128
                                newdata[col].append(round(dev,4))
                                newdata[col].append(int(splqlt+1))
                            except:
                                newdata[col].append('--')
                                newdata[col].append(int(splqlt+1))
                        else:   
                            try:
                                k107 = 0
                                for i in range(0,splqlt+1,sps):
                                    savg = np.average(data3.loc[i:i+sps-1,col])
                                    sumd = 0
                                    for j in data3.loc[i:i+sps-1,col]:
                                        sumd += (float(j)-savg)**2
                                    # print(sumd)
                                    test += sumd
                                    k107 += len(data3.loc[i:i+sps-1,col]) - 1
                                    # print(k107)
                                k107 = k107+1 
                                l108 = math.sqrt(test/(k107-1))
                                k109 = calculate_standard_deviation(k107)
                                dev = l108/k109
                                newdata[col].append(round(dev,4))
                                newdata[col].append(int(splqlt+1))
                            except:
                                newdata[col].append('--')
                                newdata[col].append(int(splqlt+1))
                        #Ca
                        try:        
                            newdata[col].append(round((2*float(rs.loc[5,col]) - float(rs.loc[1,col]) - float(rs.loc[2,col]))/(float(rs.loc[1,col]) - float(rs.loc[2,col])),4)) #ca=2avg-usl-lsl/usl-lsl
                        except:
                            newdata[col].append('--')
                        #Cp
                        
                        try:        
                            newdata[col].append(round((float(rs.loc[1,col]) - float(rs.loc[2,col]))/(6*dev),4)) #cp=(usl-lsl)/6*dev   USL, LSL != '--'
                        except:
                            try:        
                                newdata[col].append(round(abs((float(rs.loc[5,col]) - float(rs.loc[1,col])))/(3*dev),4)) # cp = |avg-usl|/3.dev USL != '--'
                            except:
                                try:
                                    newdata[col].append(round(abs((float(rs.loc[5,col]) - float(rs.loc[2,col])))/(3*dev),4)) # cp = |avg-lsl|/3.dev LSL != '--'
                                except:
                                    newdata[col].append('--') #LSL = USL = '--'

                        try:        
                            newdata[col].append(round(((float(rs.loc[1,col]) - float(rs.loc[2,col]))/(6*dev))*(1-abs((2*float(rs.loc[5,col]) - float(rs.loc[1,col]) - float(rs.loc[2,col]))/(float(rs.loc[1,col]) - float(rs.loc[2,col])))),4)) #Cpk = ((usl-lsl)/6*dev) * (1-|2avg-usl-lsl/usl-lsl|)
                        except:
                            try:        
                                newdata[col].append(round(abs((float(rs.loc[5,col]) - float(rs.loc[1,col])))/(3*dev),4)) # cpk = |avg-usl|/3.dev USL != '--'
                            except:
                                try:
                                    newdata[col].append(round(abs((float(rs.loc[5,col]) - float(rs.loc[2,col])))/(3*dev),4)) # cpk = |avg-lsl|/3.dev LSL != '--'
                                except:
                                    newdata[col].append('--') #LSL = USL = '--'
                        # P
                        count = 0
                        for x in range(splqlt):
                            try:
                                if float(data3.loc[x,col]) > float(data1.loc[1,col]):
                                    print(col, data3.loc[x,col])
                                    count += 1
                                    
                            except:
                                pass
                            try:
                                if float(data3.loc[x,col]) < float(data1.loc[2,col]):
                                    print(col, data3.loc[x,col])
                                    count += 1
                                    
                            except:
                                pass
                        print(col, count)
                        try:
                            newdata[col].append(str(round(100*count/len(data3),2))+'%') 
                        except:
                            newdata[col].append('--')
                    newdata['Control_Item'] = ['Std.Dev','Qty','Ca','Cp','Cpk','P']
                    newdata['FileID'] = [datafileid_str]*6
                    print(newdata)
                    dt1 = pd.DataFrame(newdata)
                    rs2_list = pd.concat([rs, dt1], ignore_index=True)
                    # rs3 = [None,None,None,None,None,None,None,None,None,None,None,None,None]
                    rs3 = []
                    if filegroup =='DS50000-Heat-Treatment':
                        for index, row in rs2_list.iterrows():
                            item = {
                                "Control_Item": row["Control_Item"],
                                "1_Yield Strength": row["1_Yield Strength"],
                                "2_Tensile Strength": row["2_Tensile Strength"],
                                "3_Elongation": row["3_Elongation"],
                                "4_Cross-Sectional Area": row["4_Cross-Sectional Area"],
                                "5_Hardness": row["5_Hardness"],
                                "A1_Appearance": row["A1_Appearance"]
                                
                            }
                            rs3.append(item)
                    elif filegroup =='DC10200-Casting':
                        for index, row in rs2_list.iterrows():
                            item = {
                                "Control_Item": row["Control_Item"],
                                "C": row["C"],
                                "Si": row["Si"],
                                "Mn": row["Mn"],
                                "P": row["P"],
                                "S": row["S"],
                                "Ni": row["Ni"],
                                "Cr": row["Cr"],
                                "Mo": row["Mo"],
                                "Cu": row["Cu"],
                                "Ti": row["Ti"],
                                "V": row["V"],
                                "Pb": row["Pb"],
                                "W": row["W"],
                                "Al": row["Al"],
                                "Co": row["Co"],
                                "Nb": row["Nb"],
                                "As": row["As"],
                                "Sn": row["Sn"],
                                "Sb": row["Sb"],
                                "B": row["B"],
                                "Bi": row["Bi"],
                                "Ca": row["Ca"],
                                "Zn": row["Zn"],
                                "N": row["N"],
                                "Ce": row["Ce"],
                                "Mg": row["Mg"],
                                "Ta": row["Ta"],
                                "Zr": row["Zr"],
                                "Ti+Nb": row["Ti+Nb"],
                                "Fe%": row["Fe%"],
                                "CEF": row["CEF"],
                                "Other": row["Other"],
                                "A1_Appearance": row["A1_Appearance"],
                            }
                            rs3.append(item)

                    
                    print(rs3)

                    
                
                    return render_template('pages/searchoptical.html',user=user, data = groupname,process = "OpticalDivision",post=post,post_flag = True,error_flag=True, timestart=timestart,timeend=timeend,opticaldivision_data=opticaldivision_data, rs3 = rs3)
                except Exception:
                    # Systemp_log(traceback.format_exc()).append_new_line()
                    return render_template('pages/searchoptical.html',user=user, data = groupname,process = "OpticalDivision",post=post,post_flag = True,error_flag=False, timestart=timestart,timeend=timeend,opticaldivision_data=opticaldivision_data, rs3 = None)

            else:
                conn1 = pyodbc.connect('Driver={SQL Server}; Server=192.168.9.246; uid=sa; pwd=saa;Database=SPC; Trusted_Connection=No;',timeout=1)
                # group = "select GroupName from FileGroup_V WHERE GroupName LIKE '%DC%' or GroupName Like '%DS%' or GroupName Like '%GC%' or GroupName Like '%Engineer%' or GroupName like '%VC-Test%' ORDER BY GroupName"
                group = "select GroupName from FileGroup_V WHERE GroupName ='DC10200-Casting' "
                cursor1 = conn1.cursor()   
                cursor1.execute(group)
                groupname= cursor1.fetchall()
                
                return render_template('pages/searchoptical.html',error_flag=True, data = groupname,user=user,process = "OpticalDivision")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
    
    
@app.route("/OpticalDivision3", methods=['POST', 'GET'])
def Searchoptical3():
    payload = request.get_json()
    filegroup = payload.get('grpname')   
    filename = payload.get('filename')   
    conn1 = pyodbc.connect('Driver={SQL Server}; Server=192.168.9.246; uid=sa; pwd=saa;Database=SPC; Trusted_Connection=No;',timeout=1)
    
    print(filegroup)
    fileid = "select FileID From FileName where FileGroup = '"+filegroup+"' and  FileName='"+filename+"'"
    print(fileid)
    cursor2 = conn1.cursor()
    cursor2.execute(fileid)
    fileidname= cursor2.fetchall()
    print(fileidname)
    data3 = []
    for i in fileidname:
        data3.append(i[0])
    datafileid = {
        "data":data3
    }
    session['datafileid'] = datafileid

    return jsonify(datafileid)
    
@app.route("/OpticalDivision2", methods=['POST', 'GET'])
def Searchoptical2():
    payload = request.get_json()
    filegroup = payload.get('grpname')      
    conn1 = pyodbc.connect('Driver={SQL Server}; Server=192.168.9.246; uid=sa; pwd=saa;Database=SPC; Trusted_Connection=No;',timeout=1)
    print("filegroup")
    print(filegroup)
    file = "select FileName FROM FileName where FileGroup = '"+filegroup+"'"
    print(file)
    cursor2 = conn1.cursor()
    cursor2.execute(file)
    filename= cursor2.fetchall()
    print(filename)
    data2 = []
    for i in filename:
        data2.append(i[0])
    data = {
        "data":data2
    }
    return jsonify(data)
@app.route("/lwm", methods=['POST', 'GET'])
def openlwm():
    file_list = []
    folder_path = 'static/csv'
    for f in os.listdir(folder_path):
        
        if 'csv' in f or 'CSV' in f:
            print(f)
            file_list.append(f[:-4])
    print(file_list)
    return render_template('pages/lwm.html',file_list=file_list)



@app.route("/lwm/<filename>", methods=['POST', 'GET'])
def lwm(filename):

    data = pd.read_csv('static/csv/'+filename,skiprows=8,sep=';')
    print(data)
    x_axis=[]
    y_axis=[]
    list_a=[]
    list_b=[]
    list_c=[]
    list_d=[]
    list_a1=[]
    list_b1=[]
    list_c1=[]
    list_d1=[]
    tolerance_percent = 0.2  # 20%
    jitter_percent = 40     # 40%
    upper_limit = []
    lower_limit = []
    upper_limit1 = []
    lower_limit1 = []
    upper_limit2 = []
    lower_limit2 = []
    upper_limit3 = []
    lower_limit3 = []
    count = 99999999
    for i in range(len(data)):
        x_axis.append(data.loc[i,"Time"])
        list_a.append(data.loc[i,"Plasma"])
        list_b.append(data.loc[i,"Temp"])
        list_c.append(data.loc[i,"Refl"])
        list_d.append(data.loc[i,"Laser"])
        list_a1.append(data.loc[i,"P-Raw"])
        list_b1.append(data.loc[i,"T-Raw"])
        list_c1.append(data.loc[i,"R-Raw"])
        list_d1.append(data.loc[i,"L-Raw"])
        tolerance = data.loc[i,"Plasma"] * tolerance_percent
        tolerance1 = data.loc[i,"Temp"] * tolerance_percent
        tolerance2 = data.loc[i,"Refl"] * tolerance_percent
        tolerance3 = data.loc[i,"Laser"] * tolerance_percent
        if count>jitter_percent:
            lower_bound = data.loc[i,"Plasma"] - tolerance
            upper_bound = data.loc[i,"Plasma"] + tolerance
            lower_bound1 = data.loc[i,"Temp"] - tolerance1
            upper_bound1 = data.loc[i,"Temp"] + tolerance1
            lower_bound2 = data.loc[i,"Refl"] - tolerance2
            upper_bound2 = data.loc[i,"Refl"] + tolerance2
            lower_bound3 = data.loc[i,"Laser"] - tolerance3
            upper_bound3 = data.loc[i,"Laser"] + tolerance3
            count=0
        # Thêm giới hạn vào danh sách
        lower_limit.append(lower_bound)
        upper_limit.append(upper_bound)
        lower_limit1.append(lower_bound1)
        upper_limit1.append(upper_bound1)
        lower_limit2.append(lower_bound2)
        upper_limit2.append(upper_bound2)
        lower_limit3.append(lower_bound3)
        upper_limit3.append(upper_bound3)
        count+=1
    figure,axis = plt.subplots(4,1)
    figure.tight_layout(pad=1.0)
    axis[0].plot(x_axis, list_a)
    axis[0].plot(x_axis, lower_limit)
    axis[0].plot(x_axis, upper_limit)
    
    axis[1].plot(x_axis, list_b)
    axis[1].plot(x_axis, lower_limit1)
    axis[1].plot(x_axis, upper_limit1)
    axis[2].plot(x_axis, list_c)
    axis[2].plot(x_axis, lower_limit2)
    axis[2].plot(x_axis, upper_limit2)
    axis[3].plot(x_axis, list_d)
    axis[3].plot(x_axis, lower_limit3)
    axis[3].plot(x_axis, upper_limit3)
    axis[3].set_ylim(0,3.4)
    axis[0].set_title("Plasma")
    axis[1].set_title("Temp")
    axis[2].set_title("Relf")
    axis[3].set_title("Laser")
    plt.gcf().set_size_inches(10, 5)
    plt.savefig('static/csv/'+filename[:-4]+'.png', dpi=170  )
    return '/static/csv/'+filename[:-4]+'.png'

@app.route("/Dongbin", methods=['POST', 'GET'])
def open_Dongbin():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process = "Dongbin"
        if check_secr(24): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)*  From [QC].[dbo].[DONG_BIN] where TimeFinish >'"+timestart+"' and TimeFinish <'"+timeend+"'")
                #print("SELECT Top(15000)* From[QC].[dbo].[air_tight] where Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df1 = pd.read_sql(
                    "   SELECT left(TimeFinish,11) as 'Time', Result, COUNT(Result) AS 'Count' FROM [QC].[dbo].[DONG_BIN] where  TimeFinish >'"+timestart+"' and TimeFinish<'"+timeend+"'GROUP BY left(TimeFinish,11),Result order by left(TimeFinish,11)",
                    conn)
                    plt.rc('font', size=14)
                    
                    df1 = df1.pivot_table(index='Time', columns='Result')
                    
                    ax1 = df1.plot(kind='bar', y='Count', figsize=(20, 10), color=['red','darkgreen' ], width = 0.7)
                    for container in ax1.containers:
                        ax1.bar_label(container, rotation=0)
                    ax1.legend(title= 'Result',  bbox_to_anchor= (1, 1.13), ncol = 2)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    # plt.figure(figsize=(2, 2))
                    plt.xticks(rotation=45)
                    plt.title("PIN PRESS DATA")
                    plt.savefig('static/images/dongbin.png')
                # pagecount = int(soluong/100)
                # print(data['DMCin'])
                return render_template('pages/Dongbin.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Dongbin.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Classification036", methods=['POST', 'GET'])
def open_Classification036():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Classification036"
        if check_secr(28): 
            if request.method == 'POST':
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                cursor = conn.cursor()
                # data = laser_all_data("","","",timestart,timeend)
                cursor.execute("SELECT Top(15000)*   FROM [QC].[dbo].[AirGauge036_Classification] where [TimeDMC] >'"+timestart+"' and [TimeDMC] <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                # html_data = data.to_html(classes='table table-striped',table_id="example",border=0,index=False)
                    df2 = pd.read_sql(
                        " SELECT CAST(TimeDMC AS DATE)as Time,  COUNT(DMC) AS 'Quantity' FROM [QC].[dbo].[AirGauge036_Classification] where  TimeDMC >'"+timestart+"' and TimeDMC<'"+timeend+"' GROUP BY CAST(TimeDMC AS DATE) order by CAST(TimeDMC AS DATE) ",
                        conn)
                    
                    df2['Time'] = pd.to_datetime(df2['Time'])

                    # Tạo biểu đồ cột
                    plt.figure(figsize=(20, 10))
                    ax = plt.bar(range(len(df2)), df2['Quantity'], color='darkgreen', width=0.7)


                    # Đặt nhãn và tiêu đề
                    plt.xlabel('Time')
                    plt.ylabel('Quantity')
                    
                    plt.gca().xaxis.set_major_formatter(plt.FixedFormatter(df2['Time'].dt.strftime('%Y-%m-%d')))
                    plt.gca().xaxis.set_major_locator(plt.FixedLocator(range(len(df2))))


                    plt.xticks(rotation=45)

                    # Thêm giá trị trên đỉnh cột
                    for bar in ax:
                        height = bar.get_height()
                        plt.annotate(f'{height}',
                                    xy=(bar.get_x() + bar.get_width() / 2, height),
                                    xytext=(0, 3),  # 3 points vertical offset
                                    textcoords="offset points",
                                    ha='center', va='bottom')
                    # Hiển thị biểu đồ
                    # plt.tight_layout()
                    plt.title('Classification A1811036')
                    plt.savefig('static/images/Classification036.png')
                
                return render_template('pages/Classification036.html',user=user,process=process,post_flag = True,data = data,soluong = soluong, timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Classification036.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))

@app.route("/Classification3123", methods=['POST', 'GET'])
def open_Classification3123():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Classification3123"
        if check_secr(33): 
            if request.method == 'POST':
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                cursor = conn.cursor()
                # data = laser_all_data("","","",timestart,timeend)
                cursor.execute("SELECT Top(15000) *   FROM [QC].[dbo].[Classification_A2303123] where [TimeSave] >='"+timestart+"' and [TimeSave] <='"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                # html_data = data.to_html(classes='table table-striped',table_id="example",border=0,index=False)
                    df2 = pd.read_sql(
                        " SELECT CAST(TimeSave AS DATE)as Time,  COUNT(DMC_Scan) AS 'Quantity' FROM [QC].[dbo].[Classification_A2303123] where  TimeSave >='"+timestart+"' and TimeSave<='"+timeend+"' GROUP BY CAST(TimeSave AS DATE) order by CAST(TimeSave AS DATE) ",
                        conn)
                    
                    df2['Time'] = pd.to_datetime(df2['Time'])

                    # Tạo biểu đồ cột
                    plt.figure(figsize=(20, 10))
                    ax = plt.bar(range(len(df2)), df2['Quantity'], color='darkgreen', width=0.7)


                    # Đặt nhãn và tiêu đề
                    plt.xlabel('Time')
                    plt.ylabel('Quantity')
                    plt.title('Classification A2303123')
                    plt.gca().xaxis.set_major_formatter(plt.FixedFormatter(df2['Time'].dt.strftime('%Y-%m-%d')))
                    plt.gca().xaxis.set_major_locator(plt.FixedLocator(range(len(df2))))


                    plt.xticks(rotation=45)

                    # Thêm giá trị trên đỉnh cột
                    for bar in ax:
                        height = bar.get_height()
                        plt.annotate(f'{height}',
                                    xy=(bar.get_x() + bar.get_width() / 2, height),
                                    xytext=(0, 3),  # 3 points vertical offset
                                    textcoords="offset points",
                                    ha='center', va='bottom')
                    # Hiển thị biểu đồ
                    # plt.tight_layout()
                    # plt.title("Classification 036")
                    plt.savefig('static/images/Classification3123.png')
                
                return render_template('pages/Classification3123.html',user=user,process=process,post_flag = True,data = data,soluong = soluong, timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Classification3123.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route('/download/Classification3123',methods=['POST'])
def download_Classification3123():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df8 = pd.read_sql("SELECT * FROM [QC].[dbo].[Classification_A2303123] where TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df8.to_excel(writer, sheet_name='Classification3123',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Classification3123_Data.xls"})
@app.route('/download/Classification036',methods=['POST'])
def download_Classification036():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df8 = pd.read_sql("SELECT * FROM [QC].[dbo].[AirGauge036_Classification] where [TimeDMC] >'"+timestart+"' and [TimeDMC] <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df8.to_excel(writer, sheet_name='Classification036',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Classification036_Data.xls"})
@app.route('/download/Dongbin',methods=['POST'])
def download_DONGBIN():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    print(timestart)
    df7 = pd.read_sql("SELECT * From [QC].[dbo].[DONG_BIN] where TimeFinish >'"+timestart+"' and TimeFinish <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df7.to_excel(writer, sheet_name='PINPRESS',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=PINPRESS_data.xls"})
@app.route("/Laser", methods=['POST', 'GET'])
def open_laser():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Laser"
        if check_secr(14): 
            if request.method == 'POST':
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                cursor = conn.cursor()
                # data = laser_all_data("","","",timestart,timeend)
                cursor.execute("SELECT Top(15000)* From[QC].[dbo].[Laser] where TimeOutBarcode >'"+timestart+"' and TimeOutBarcode <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                # html_data = data.to_html(classes='table table-striped',table_id="example",border=0,index=False)
                    df2 = pd.read_sql(
                        " SELECT left(TimeInDMC,11)as Time, Quality, COUNT(Quality) AS 'Count' FROM [QC].[dbo].[Laser] where  TimeInDMC >'"+timestart+"' and TimeInDMC<'"+timeend+"'GROUP BY left(TimeInDMC,11),Quality order by left(TimeInDMC,11)",
                        conn)
                    
                    plt.rc('font', size=13)
                    df2 = df2.pivot_table(index='Time', columns='Quality')
                    ax2 = df2.plot(kind='bar', figsize=(20, 10), color=['darkgreen','red' ,'limegreen', 'yellow', 'orange', 'purple'], width = 0.7)
                    for container in ax2.containers:
                        ax2.bar_label(container, rotation=0)
                    ax2.legend(title= 'Quality',  bbox_to_anchor= (1, 1.215), ncol = 2)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    # plt.figure(figsize=(2, 2))
                    plt.xticks(rotation=45)
                    plt.title("A2012003/004-DMC QUALITY OF CASTING AFTER LASER MARKING")
                    plt.savefig('static/images/Laser.png')
                
                return render_template('pages/Laser.html',user=user,process=process,post_flag = True,data = data,soluong = soluong, timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Laser.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Airgauge036", methods=['POST', 'GET'])
def open_Airgauge036():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Airgauge036"
        if check_secr(14): 
            if request.method == 'POST':
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                cursor = conn.cursor()
                # data = laser_all_data("","","",timestart,timeend)
                cursor.execute("SELECT Top(15000)*   FROM [QC].[dbo].[AirGauge036] where TimeFinish >'"+timestart+"' and TimeFinish <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                # html_data = data.to_html(classes='table table-striped',table_id="example",border=0,index=False)
                    df2 = pd.read_sql(
                        " SELECT left(TimeFinish,11)as Time, Result, COUNT(Result) AS 'Count' FROM [QC].[dbo].[AirGauge036] where  TimeFinish >'"+timestart+"' and TimeFinish<'"+timeend+"'GROUP BY left(TimeFinish,11), Result order by left(TimeFinish,11)",
                        conn)
                    
                    plt.rc('font', size=13)
                    df2 = df2.pivot_table(index='Time', columns='Result')
                    ax2 = df2.plot(kind='bar', figsize=(20, 10), color=['red', 'darkgreen','yellow', 'limegreen'], width = 0.7)
                    for container in ax2.containers:
                        ax2.bar_label(container, rotation=0)
                    ax2.legend(title= 'Quality',  bbox_to_anchor= (1, 1.215), ncol = 2)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    # plt.figure(figsize=(2, 2))
                    plt.xticks(rotation=45)
                    plt.title("AIR GAUGE 036")
                    plt.savefig('static/images/airgauge036.png')
                
                return render_template('pages/Airgauge036.html',user=user,process=process,post_flag = True,data = data,soluong = soluong, timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Airgauge036.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route('/download/Airgauge036',methods=['POST'])
def download_Airgauge036():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df8 = pd.read_sql("SELECT * FROM [QC].[dbo].[AirGauge036] where [TimeFinish] >'"+timestart+"' and [TimeFinish] <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df8.to_excel(writer, sheet_name='Airgauge036',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Airgauge036_Data.xls"})
@app.route('/download/CMM/Sixpack004',methods=['POST'])
def download_Sixpack004():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    cmd = ["No.13 Flatness","No.16 Flatness","No.29 Flatness - qc","No.33 Diameter - qc","No.94 Distance - sc","No.95 Position - sc","No.96 Distance - sc","No.97 Position - sc","No.114 Position - qc","No.126 Flatness - qc","No.130 Diameter - qc"]
    name = ["No.13 Flatness","No.16 Flatness","No.29 Flatness - qc","No.33 Diameter - qc","No.94 Distance - sc","No.95 Position - sc","No.96 Distance - sc","No.97 Position - sc","No.114 Position - qc","No.126 Flatness - qc","No.130 Diameter - qc"]
    for cmmid in cmd:
        data2 = pd.read_sql("SELECT  [DMC],[TimeSave],[Product],[Line] ,[CMMCode],[actual],[nominal],[uppertol],[lowertol],[deviation] ,[Result] FROM [QC].[dbo].[CMMdata] where id = '"+cmmid+"' and Product like '%BW A2012004%' and TimeSave>'"+timestart+"' and TimeSave<'"+timeend+"'  order by TimeSave",conn)

        
        data2.to_excel(writer, sheet_name=cmmid,index=False)
    
        # Tạo một đối tượng xlsxwriter.Workbook để lấy tệp excel và sheet cụ thể
        # workbook = writer.book
        # worksheet = writer.sheets[cmmid]
        
        # # Thêm hình ảnh vào cuối sheet
        # worksheet.insert_image('E10', "/static/images/"+cmmid+'.PNG', {'x_offset': 15, 'y_offset': 10, 'x_scale': 0.5, 'y_scale': 0.5})
    
    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=CMM_Sixpack004.xls"})
@app.route("/SPC/Sixpack", methods=['POST', 'GET'])
def SPCSixpack1():
    
    # global datafileid_str, rs2_list
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(4): 
            if request.method == 'POST':
                # DMC=request.form['DMC']   
                # Castingcode = request.form['Castingcode']
                # Productcode = request.form['Productcode']
                
                # error_flag = False
                timestart=request.form['daystart']+" "+request.form['timestart']
                if request.form['daystart'] == '':
                    timestart = '1900-01-01' + timestart
                timeend=request.form['dayend']+" "+request.form['timeend']
                if request.form['dayend'] == '':
                    timeend = '2100-01-01'+ timeend
                filegroup=request.form['grpname']
                filename=request.form['filename']
                # print('aaaaaaa',filegroup)
                # print('bbbbbbb',filename)
                

                conn1 = pyodbc.connect('Driver={SQL Server}; Server=192.168.9.246; uid=sa; pwd=saa;Database=SPC; Trusted_Connection=No;',timeout=1)
                # group = "select GroupName from FileGroup_V WHERE GroupName LIKE '%DC%' or GroupName Like '%DS%' or GroupName Like '%GC%' or GroupName Like '%Engineer%' or GroupName like '%VC-Test%' ORDER BY GroupName"
                group = "select GroupName from FileGroup_V WHERE GroupName ='DC10200-Casting' "
                cursor1 = conn1.cursor()   
                cursor1.execute(group)
                groupname= cursor1.fetchall()
                datafileid = session.get('datafileid')
                # print('fileid0',datafileid)
                datafileid_str = str(datafileid["data"][0])
                # print('fileid1',datafileid_str)
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=SPC; Trusted_Connection=No;',timeout=1)
                cursor = conn.cursor()            
                

                
                return render_template('pages/SPCsixpack1.html',user=user, data = groupname,process = "SPC",post_flag = True,error_flag=True, timestart=timestart,timeend=timeend)
                
            else:
                conn1 = pyodbc.connect('Driver={SQL Server}; Server=192.168.9.246; uid=sa; pwd=saa;Database=SPC; Trusted_Connection=No;',timeout=1)
                # group = "select GroupName from FileGroup_V WHERE GroupName LIKE '%DC%' or GroupName Like '%DS%' or GroupName Like '%GC%' or GroupName Like '%Engineer%' or GroupName like '%VC-Test%' ORDER BY GroupName"
                group = "select GroupName from FileGroup_V WHERE GroupName ='DC10200-Casting' "
                cursor1 = conn1.cursor()   
                cursor1.execute(group)
                groupname= cursor1.fetchall()
                
                return render_template('pages/SPCsixpack1.html',error_flag=True, data = groupname,user=user,process = "OpticalDivision")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))

@app.route("/getSPCsixpack", methods=['POST', 'GET'])
def getSPCsixpack():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("SELECT distinct(COALESCE(SUBSTRING([LevelItem], 1, CHARINDEX(';', [LevelItem] + ';') - 1), [LevelItem])) AS Product FROM [SPC].[dbo].[FileData]",conn)
    edata = {}
    edata["Product"]=[]
    for i in range(len(data)):
        edata["Product"].append(data.loc[i,"Product"])
    return jsonify(edata)    
@app.route("/getSPCsixpackform/<grpname>/<filename>", methods=['GET'])
def getSPCsixpackform(grpname,filename):
    print('111',grpname,filename)
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=SPC; Trusted_Connection=No;',timeout=1)
    cursor = conn.cursor()
    cursor.execute("SELECT distinct(CTRLITEM) as CTRLITEM FROM [SPC].[dbo].[VCTRL] where [FileID]= (select FileID From FileName where FileGroup = '"+grpname+"' and  FileName='"+filename+"')")

    # cursor.execute("select  [CTRLITEM],product.nameproduct from [SPC].[dbo].[VCTRL] as name  left join ( SELECT [FileID], COALESCE(SUBSTRING([LevelItem], 1, CHARINDEX(';', [LevelItem] + ';') - 1), [LevelItem]) AS nameproduct FROM [SPC].[dbo].[FileData]  group by [FileID], COALESCE(SUBSTRING([LevelItem], 1, CHARINDEX(';', [LevelItem] + ';') - 1), [LevelItem]) ) as product on name.[FileID] = product.[FileID] where product.nameproduct ='"+mahang+"' group by [CTRLITEM],product.nameproduct order by CTRLITEM asc")
    # print("select  [CTRLITEM],product.nameproduct  from [SPC].[dbo].[VCTRL] as name  left join ( SELECT [FileID], COALESCE(SUBSTRING([LevelItem], 1, CHARINDEX(';', [LevelItem] + ';') - 1), [LevelItem]) AS nameproduct FROM [SPC].[dbo].[FileData]  group by [FileID], COALESCE(SUBSTRING([LevelItem], 1, CHARINDEX(';', [LevelItem] + ';') - 1), [LevelItem]) ) as product on name.[FileID] = product.[FileID] where product.nameproduct ='"+mahang+"' group by [CTRLITEM],product.nameproduct order by CTRLITEM asc")
    data = cursor.fetchall()
    
    edata = []
    for dt in data:
        edata.append({
            "CMM":dt[0],
            "Name":dt[0]
        })
    return jsonify(edata)
@app.route("/createSPCsixpackimg", methods=['POST'])
def createSPCsixpackimg():
    data = request.json
    print('456',data)
    grpname = data['grpname']
    filename = data['filename']
    dulieu = data['name']
    name = data['name']
    timestart = data['daystart'] + ' ' + data['hourstart'] + ':' + data['minutestart']
    timeend = data['dayend'] + ' ' + data['hourend'] + ':' + data['minuteend']
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=SPC; Trusted_Connection=No;',timeout=1)
    #url = 'http://192.168.8.21:5010/api_post_json_data'
    url = 'http://192.168.8.65:5009/six_pack_v2_1'
    # cmd = "SELECT  [actual],[nominal],[uppertol],[lowertol] FROM [QC].[dbo].[CMMdata] where id = '"+dulieu+"' and Product like '%"+mahang+"%' and TimeSave>'"+timestart+"' and TimeSave<'"+timeend+"'  order by TimeSave"
    data2 = pd.read_sql("select VCDATA.VCTRLID, Name,Value, SC, SU, SL from VCDATA left join displayname on VCDATA.Linkname = Displayname.ID left join (SELECT *  FROM [SPC].[dbo].[VCTRL] where [FileID]= (select FileID From FileName where FileGroup = '"+grpname+"' and  FileName='"+filename+"')) as a on Displayname.Name = a.CTRLITEM   where VCDATA.VCTRLID in (select [VCTRLID] FROM [SPC].[dbo].[FileData] where DateTimeCreate >='"+timestart+"' and DateTimeCreate <='"+timeend+"' and [FileID]=(select FileID From FileName where FileGroup = '"+grpname+"' and  FileName='"+filename+"')) and Linkname != 0 and Name = '"+dulieu+"'order by VCDATA.VCTRLID",conn)
    # data2 = pd.read_sql("select * from (select [FileData].VCTRLID, COALESCE(SUBSTRING([LevelItem], 1, CHARINDEX(';', [LevelItem] + ';') - 1), [LevelItem]) AS Product, DateTimeCreate,[VCDATA].Value,Displayname.Name, [VCTRL].SC,[VCTRL].SU, [VCTRL].SL from [SPC].[dbo].[FileData] left join [VCDATA] on [FileData].VCTRLID = [VCDATA].VCTRLID left join displayname on VCDATA.Linkname = Displayname.ID left join [VCTRL] on [FileData].VCTRLID = [VCTRL].VCTRLID) as abcd where DateTimeCreate >='"+timestart+"' and DateTimeCreate <='"+timeend+"' and Name = '"+dulieu+"'  and Product like '%"+mahang+"%' order by DateTimeCreate asc",conn)

    # data2 = pd.read_sql("SELECT  [actual],[nominal],[uppertol],[lowertol] FROM [QC].[dbo].[CMMdata] where id = '"+dulieu+"' and Product like '%"+mahang+"%' and TimeSave>'"+timestart+"' and TimeSave<'"+timeend+"'  order by TimeSave",conn)
    print("select VCDATA.VCTRLID, Name,Value, SC, SU, SL from VCDATA left join displayname on VCDATA.Linkname = Displayname.ID left join (SELECT *  FROM [SPC].[dbo].[VCTRL] where [FileID]= (select FileID From FileName where FileGroup = '"+grpname+"' and  FileName='"+filename+"')) as a on Displayname.Name = a.CTRLITEM   where VCDATA.VCTRLID in (select [VCTRLID] FROM [SPC].[dbo].[FileData] where DateTimeCreate >='"+timestart+"' and DateTimeCreate <='"+timeend+"' and [FileID]=(select FileID From FileName where FileGroup = '"+grpname+"' and  FileName='"+filename+"')) and Linkname != 0 and Name = '"+dulieu+"'order by VCDATA.VCTRLID")
    lst = []
    ust = []
    for x in data2["SL"]:
        try:
            lst.append(float(x))
            # lst.append(x)
        except:
            pass
    for x in data2["SU"]:
        try:
            ust.append(float(x))
            # ust.append(x)
        except:
            pass
    print(lst)
    print(ust)
    
    if len(lst) == 0:
        lsl = None
    else:
        print(sum(lst),len(lst))
        lsl = sum(lst) / len(lst) 
    if len(ust) == 0:
        usl = None
    else:
        print(sum(ust),len(ust))
        usl = sum(ust) / len(ust)
    data = {
    "LSL":  lsl,
    "USL": usl,
    "data": [float(x) for x in data2["Value"].tolist()],
    "name": name
    }
    print(len(data),data)
    if len(data["data"]) > 0:
        Req = requests.post(url=url,json=data)
        # print(Req.text)
        rs = '<style></style><div id="fig_el2014026492006797128259275921"></div><script>console.log("abc");</script>'
        return Response(Req.text, content_type='text/html')
    else:
        return 'No Data'

@app.route("/CMM/Sixpack", methods=['POST', 'GET'])
def open_Sixpack():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="CMM"
        if check_secr(5): 
            return render_template('/pages/CMMsixpack.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/getProductsixpack", methods=['POST', 'GET'])
def getProductsixpack():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("select distinct(Product) as Product from CMMsixpackform  order by Product",conn)
    edata = {}
    edata["Product"]=[]
    for i in range(len(data)):
        edata["Product"].append(data.loc[i,"Product"])
    return jsonify(edata)




@app.route("/createsixpack", methods=['POST'])
def createsixpack():
    data = request.json
    mahang = data['mahang']
    dulieu = data['data']
    print(mahang)
    print(dulieu)
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    cursor = conn.cursor()
    cursor.execute("delete FROM [QC].[dbo].[CMMsixpackform] where product = '"+mahang+"'")
    for dt in dulieu:
        cursor.execute("insert into [QC].[dbo].[CMMsixpackform] values (?,?,?)",
                       mahang,
                       dt["CMM"],
                       dt["Name"])
    cursor.commit()
    conn.close()
    return "OK"
@app.route("/CMM/Sixpackform", methods=['GET'])
def Sixpackform():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="CMM"
        if check_secr(8): 
            return render_template("/pages/sixpackform.html")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
    
@app.route("/createsixpackimg", methods=['POST'])
def createsixpackimg():
    data = request.json
    mahang = data['mahang']
    dulieu = data['data']
    name = data['name']
    timestart = data['daystart'] + ' ' + data['hourstart'] + ':' + data['minutestart']
    timeend = data['dayend'] + ' ' + data['hourend'] + ':' + data['minuteend']
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    #url = 'http://192.168.8.21:5010/api_post_json_data'
    url = 'http://192.168.8.65:5009/six_pack_v2_1'
    cmd = "SELECT  [actual],[nominal],[uppertol],[lowertol] FROM [QC].[dbo].[CMMdata] where id = '"+dulieu+"' and Product like '%"+mahang+"%' and TimeSave>'"+timestart+"' and TimeSave<'"+timeend+"'  order by TimeSave"
    data2 = pd.read_sql("SELECT  [actual],[nominal],[uppertol],[lowertol] FROM [QC].[dbo].[CMMdata] where id = '"+dulieu+"' and Product like '%"+mahang+"%' and TimeSave>'"+timestart+"' and TimeSave<'"+timeend+"'  order by TimeSave",conn)
    lst = []
    ust = []
    for x in data2["lowertol"]:
        try:
            lst.append(float(x))
        except:
            pass
    for x in data2["uppertol"]:
        try:
            ust.append(float(x))
        except:
            pass
    if len(lst) == 0:
        lsl = None
    else:
        lsl = sum(lst) / len(lst)+float(data2.loc[0,"nominal"])
    if len(ust) == 0:
        usl = None
    else:
        usl = sum(ust) / len(ust)+float(data2.loc[0,"nominal"])
    data = {
    "LSL":  lsl,
    "USL": usl,
    "data": [float(x) for x in data2["actual"].tolist()],
    "name": name
    }
    print(data)
    if len(data["data"]) > 0:
        Req = requests.post(url=url,json=data)
        print(Req.text)
        rs = '<style></style><div id="fig_el2014026492006797128259275921"></div><script>console.log("abc");</script>'
        return Response(Req.text, content_type='text/html')
    else:
        return 'No Data'
    # return Response(Req.content, mimetype='image/png')
@app.route("/getsixpackform/<mahang>", methods=['GET'])
def getsixpackform(mahang):
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    cursor = conn.cursor()
    cursor.execute("select * FROM [QC].[dbo].[CMMsixpackform] where product = '"+mahang+"'  order by name")
    data = cursor.fetchall()
    edata = []
    for dt in data:
        edata.append({
            "CMM":dt[1],
            "Name":dt[2]
        })
    return jsonify(edata)
@app.route("/getProduct/<product>", methods=['POST', 'GET'])
def getProduct2(product):
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("select distinct(id) as Product from CMMData where Product like '%"+product+"%' order by id",conn)
    edata = {}
    edata["Product"]=[]
    for i in range(len(data)):
        edata["Product"].append(data.loc[i,"Product"])
    print(edata)
    # data = data.to_json()
    return jsonify(edata)
@app.route("/AirtightChamfer", methods=['POST', 'GET'])
def open_AirtightChamfer():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process = "AirtightChamfer"
        if check_secr(22): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                cursor = conn.cursor()
                cursor.execute("SELECT TOP 15000 * FROM  QC.dbo.air_tight_chamfer  WHERE Time_Start >='"+timestart+"' and Time_Start <='"+timeend+"'")
                #print("SELECT Top(15000)* From[QC].[dbo].[air_tight] where Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df1 = pd.read_sql(
                    "   SELECT left(Time_Start,11) as 'Time', Quality, COUNT(Quality) AS 'Count' FROM  QC.dbo.air_tight_chamfer  WHERE Time_Start >='"+timestart+"' and Time_Start <='"+timeend+"' GROUP BY left(Time_Start,11), Quality  order by left(Time_Start,11) desc",
                    conn)
                    plt.rc('font', size=14)
                    
                    df1 = df1.pivot_table(index='Time', columns='Quality')
                    
                    ax1 = df1.plot(kind='bar', y='Count', figsize=(20, 10), color=['red', 'darkgreen'], width = 0.7)
                    for container in ax1.containers:
                        ax1.bar_label(container, rotation=0)
                    ax1.legend(title= 'Quality',  bbox_to_anchor= (1, 1.13), ncol = 2)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    # plt.figure(figsize=(2, 2))
                    plt.xticks(rotation=45)
                    plt.title("A2012003/004-AIR LEAKAGE TEST CHAMFER")
                    plt.savefig('static/images/AIR_LEAKAGE_chamfer.png')
                # pagecount = int(soluong/100)
                # print(data['DMCin'])
                return render_template('pages/AirtightChamfer.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/AirtightChamfer.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))        
@app.route("/Airtightwindow", methods=['POST', 'GET'])
def open_Airtightwindow():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process = "Airtightwindow"
        if check_secr(22): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                cursor = conn.cursor()
                cursor.execute("SELECT TOP 15000 * FROM  QC.dbo.[air_tight_window]  WHERE Time_Finish_2 >='"+timestart+"' and Time_Finish_2 <='"+timeend+"'")
                #print("SELECT Top(15000)* From[QC].[dbo].[air_tight] where Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df1 = pd.read_sql(
                    "   SELECT left(Time_Finish_2,11) as 'Time', Status_Position_2 as Quality, COUNT(Status_Position_2) AS 'Count' FROM  QC.dbo.[air_tight_window]  WHERE Time_Finish_2 >='"+timestart+"' and Time_Finish_2 <='"+timeend+"' GROUP BY left(Time_Finish_2,11), Status_Position_2  order by left(Time_Finish_2,11) desc",
                    conn)
                    plt.rc('font', size=14)
                    quality_color_mapping = {
                        ('NG'): 'darkred',
                        ('OK'): 'darkgreen',
                        
                    }
                    df1 = df1.pivot_table(index='Time', columns='Quality')
                    
                    ax1 = df1.plot(kind='bar', y='Count', figsize=(20, 10), color=quality_color_mapping, width = 0.7)
                    for container in ax1.containers:
                        ax1.bar_label(container, rotation=0)
                    ax1.legend(title= 'Quality',  bbox_to_anchor= (1, 1.13), ncol = 2)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    # plt.figure(figsize=(2, 2))
                    plt.xticks(rotation=45)
                    plt.title("A2012004-AIR LEAKAGE TEST WINDOW")
                    plt.savefig('static/images/AIR_LEAKAGE_WINDOW.png')
                # pagecount = int(soluong/100)
                # print(data['DMCin'])
                return render_template('pages/Airtightwindow.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Airtightwindow.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/CNC", methods=['POST', 'GET'])
def open_CNC():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="CNC"
        if check_secr(17): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                datatam = pd.read_sql("SELECT MIN(id) AS MinId, MAX(id) AS MaxId FROM qc.dbo.cnc WHERE  TimeoutCNC2 >= '"+timestart+"' and TimeoutCNC2 <= '"+timeend+"'",conn)
                idmin = str(datatam.loc[0,'MinId'])
                idmax = str(datatam.loc[0,'MaxId'])
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)*  From[QC].[dbo].[CNC] where id >='"+idmin+"' and id <='"+idmax+"'")
                
                # cursor = conn.cursor()
                # cursor.execute("SELECT Top(15000)*  From[QC].[dbo].[CNC] where TimeoutCNC1 >='"+timestart+"' and TimeoutCNC1 <='"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df3 = pd.read_sql(
                        "SELECT left(TimeinCNC1, 11) as Time,[Machineno] as SetNo,sum(case when [Position1] is not null then 1 else 0 end) as Qty_OP1,sum(case when [Position2] is not null then 1 else 0 end) as Qty_OP2 FROM [QC].[dbo].[CNC] where  TimeinCNC1 >='"+timestart+"' and TimeinCNC1 <= '"+timeend+"' group by [Machineno], left(TimeinCNC1, 11)",
                        conn)
                    plt.rc('font', size=10)
                    df3['SetNo']=df3['SetNo'].apply(lambda x:int(x[3:]))
                    df3 = df3.pivot_table(index='Time', columns='SetNo')
                    ax3 = df3.plot(kind='bar', y='Qty_OP1', figsize=(20, 10), width = 0.8, color=['navy', 'darkgreen', 'salmon', 'gold', 'limegreen', 'yellow', 'orange', 'purple', 'grey', 'violet', 'peru', 'cyan','steelblue'])
                    for container in ax3.containers:
                        ax3.bar_label(container, rotation=45)
                    ax3.legend(title= 'SetNo',  bbox_to_anchor= (1, 1.17), ncol = 5)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    plt.xticks(rotation=45)
                    plt.title("OP1_Quantity")
                    plt.savefig('static/images/CNC.png')
                    df5 = pd.read_sql(
                        "SELECT left(TimeinCNC1, 11) as Time,[Machineno] as SetNo,sum(case when [Position1] is not null then 1 else 0 end) as Qty_OP1,sum(case when [Position2] is not null then 1 else 0 end) as Qty_OP2 FROM [QC].[dbo].[CNC] where  TimeinCNC1 >='"+timestart+"' and TimeinCNC1 <= '"+timeend+"' group by [Machineno], left(TimeinCNC1, 11)",
                        conn)
                    plt.rc('font', size=10)
                    df5['SetNo']=df5['SetNo'].apply(lambda x:int(x[3:]))
                    df5 = df5.pivot_table(index='Time', columns='SetNo')
                    ax5 = df5.plot(kind='bar', y='Qty_OP2', figsize=(20, 10), width = 0.8, color=['navy', 'darkgreen', 'salmon', 'gold', 'limegreen', 'yellow', 'orange', 'purple', 'grey', 'violet', 'peru', 'cyan','steelblue'])
                    for container in ax5.containers:
                        ax5.bar_label(container, rotation=45)
                    ax5.legend(title= 'SetNo',  bbox_to_anchor= (1, 1.17), ncol = 5)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    plt.xticks(rotation=45)
                    plt.title("OP2_Quantity")
                    plt.savefig('static/images/CNCop2.png')
                # pagecount = int(soluong/100)
                # print(data['DMCin'])
                return render_template('pages/CNC.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/CNC.html',user=user,process=process,timestart="",timeend="")  
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Airtight", methods=['POST', 'GET'])
def open_Airtight():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process = "Airtight"
        if check_secr(22): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                cursor = conn.cursor()
                cursor.execute("SELECT TOP 15000 * FROM  QC.dbo.air_tight  WHERE Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'")                
                #print("SELECT Top(15000)* From[QC].[dbo].[air_tight] where Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df1 = pd.read_sql(
                    "   SELECT left(Time_Start,11) as 'Time', Quality, COUNT(Quality) AS 'Count' FROM  QC.dbo.air_tight  WHERE Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"' GROUP BY left(Time_Start,11), Quality  order by left(Time_Start,11) desc",
                    conn)
                    plt.rc('font', size=14)
                    
                    df1 = df1.pivot_table(index='Time', columns='Quality')
                    
                    ax1 = df1.plot(kind='bar', y='Count', figsize=(20, 10), color=['red', 'darkgreen'], width = 0.7)
                    for container in ax1.containers:
                        ax1.bar_label(container, rotation=0)
                    ax1.legend(title= 'Quality',  bbox_to_anchor= (1, 1.13), ncol = 2)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    # plt.figure(figsize=(2, 2))
                    plt.xticks(rotation=45)
                    plt.title("A2012003/004-AIR LEAKAGE")
                    plt.savefig('static/images/AIR_LEAKAGE.png')
                # pagecount = int(soluong/100)
                # print(data['DMCin'])
                return render_template('pages/Airtight.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Airtight.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Classification", methods=['POST', 'GET'])
def open_Classification():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Classification"
        if check_secr(23): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000) [Pallet].[SPLR_LOT_NO], [Classification].* From [Classification] left join [Pallet] on [Classification].pallet_num= [Pallet].Pallet_Name where [Classification].Time_check >'"+timestart+"' and [Classification].Time_check <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df = pd.read_sql(
                    " SELECT left(Time_check,11) as 'Time', dataman_total_quality, COUNT(dataman_total_quality) AS 'Count' FROM [QC].[dbo].[Classification] where Time_check >'"+timestart+"' and Time_check <'"+timeend+"' GROUP BY left(Time_check,11),dataman_total_quality  order by left(Time_check,11) desc",
                    conn)
                    
                    plt.rc('font', size=14)
                    df = df.pivot_table(index='Time', columns='dataman_total_quality')
                    ax = df.plot(kind='bar',  figsize=(20, 10), color=['red', 'darkgreen','limegreen', 'yellow', 'orange', 'purple'], width = 0.7)
                    
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    # plt.legend(title = "Quality")
                    # ax.legend(bbox_to_anchor=(0.9, 0.9))
                    ax.legend(title= 'Quality',  bbox_to_anchor= (1, 1.2), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title("A2012003/2004-DMC QUALITY AFTER MACHINING BEFORE PACKING")
                    plt.savefig('static/images/CCD.png')
                # pagecount = int(soluong/100)
                # print(data['DMCin'])
                return render_template('pages/Classification.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Classification.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Coating", methods=['POST', 'GET'])
def open_Coating():
    global coatname
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Coating"
        if check_secr(12): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=Auto; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                coatname = request.form['coatingpos']
                print(coatname)
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)* FROM [Auto].[dbo].[Product_Quantity_W5_"+coatname+"] where DateTime >'"+timestart+"' and DateTime <'"+timeend+"'")
                data =cursor.fetchall()
                cursor1 = conn.cursor()
                cursor1.execute("SELECT Top(15000)* FROM [Auto].[dbo].[Product_Quantity_W5_"+coatname+"] where DateTime >'"+timestart+"' and DateTime <'"+timeend+"'")
                data1 =cursor1.fetchall()
                soluong = len(data)
                soluong1 = len(data1)
                # if soluong1==0 or soluong1>0:
                df1 = pd.read_sql("select Datetime, average_temp1, average_humid1 from data_"+coatname+"_W5 where Datetime >'"+timestart+"' and Datetime <'"+timeend+"' order by Datetime desc",conn)
                print('abc:',"select Datetime, average_temp1, average_humid1 from data_"+coatname+"_W5 where Datetime >'"+timestart+"' and Datetime <'"+timeend+"' order by Datetime desc")
                df1["Datetime"] = pd.to_datetime(df1["Datetime"], format='%Y-%m-%d %H:%M:%S.%f')
                df1 = df1.set_index('Datetime')
                target = pd.read_sql("select * from [Auto].[dbo].[Settinglimitcoatingx5] where Coating = '"+coatname+"' ",conn)
                temp_min = target.loc[0,'Tempmin']  # Thay thế bằng giá trị nhiệt độ mục tiêu của bạn
                temp_max = target.loc[0,'Tempmax']  # Thay thế bằng giá trị độ ẩm mục tiêu của bạn
                humid_min= target.loc[0,'Humidmin']
                humid_max= target.loc[0,'Humidmax']
                plt.figure(figsize=(10, 6))
                plt.scatter(df1.index, df1['average_temp1'], label='Average Temperature',color='orange',  s=5)
                plt.scatter(df1.index, df1['average_humid1'], label='Average Humidity',color='blue',  s=5)

                plt.axhline(y=temp_min, color='red', linestyle='-', label='Temp Min')
                plt.axhline(y=temp_max, color='green', linestyle='-', label='Temp Max')
                plt.axhline(y=humid_min, color='yellow', linestyle='-', label='Humidity Min')
                plt.axhline(y=humid_max, color='purple', linestyle='-', label='Humidity Max')

                plt.text(df1.index[0], temp_min, f'          {int(temp_min)}', color='r', va='bottom', ha='left')
                plt.text(df1.index[0], temp_max, f'          {int(temp_max)}', color='g', va='bottom', ha='left')
                plt.text(df1.index[0], humid_min, f'          {int(humid_min)}', color='yellow', va='bottom', ha='left')
                plt.text(df1.index[0], humid_max, f'          {int(humid_max)}', color='purple', va='bottom', ha='left')


                plt.xlabel('Datetime')
                plt.ylabel('Values')
                plt.title(coatname+' Average Temperature & Humidity',loc='left')
                plt.legend(bbox_to_anchor= (1, 1.1), ncol = 3,fontsize= 6)
                plt.xticks(rotation=45, fontsize= 8)
                plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d %H:%M'))  
                plt.tight_layout()
                plt.savefig('static/images/Coating1.png')
                #plt.show()
                return render_template('pages/Coating.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,soluong1=soluong1,timestart=timestart,timeend=timeend,coatname=coatname)
            else:
                return render_template('/pages/Coating.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)  
    else:
        return redirect(url_for('login'))
    
@app.route("/Manualscan", methods=['POST', 'GET'])
def Manualscan():
    global areascan
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Manualscan"
        if check_secr(12): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                areascan = request.form['areapos']
                cursor = conn.cursor()
                if areascan == 'All':
                    cursor.execute("select * FROM [QC].[dbo].[Scan_Repair_Data] where TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"' order by TimeSave asc")
                else:
                
                    cursor.execute("select * FROM [QC].[dbo].[Scan_Repair_Data] where KhuVuc like N'%"+areascan+"%' and TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"' order by TimeSave asc")
                print("select * FROM [QC].[dbo].[Scan_Repair_Data] where KhuVuc like N'%"+areascan+"%' and TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"' order by TimeSave asc")
                data =cursor.fetchall()
                
                soluong = len(data)
                if soluong >0:
                    quality_color_mapping = {
                        ('Gá Tổng Hợp'): 'grey',
                        ('Kích Thước Khác'): 'darkgreen',
                        ('Ngoại Quan'): 'darkblue',
                        ('Pin Ren'): 'gold',
                        ('Ren Tay'): 'orange',
                        
                    }
                    if areascan == 'All':
                        df7 = pd.read_sql("SELECT left(TimeSave,11) as 'Time',KhuVuc,  COUNT(DMC) AS 'Count'FROM [QC].[dbo].[Scan_Repair_Data] where TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"'  GROUP BY left(TimeSave,11),KhuVuc  order by left(TimeSave,11) asc",conn)
                    else:
                        df7 = pd.read_sql("SELECT left(TimeSave,11) as 'Time',KhuVuc,  COUNT(DMC) AS 'Count'FROM [QC].[dbo].[Scan_Repair_Data] where KhuVuc like N'%"+areascan+"%' and TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"'  GROUP BY left(TimeSave,11),KhuVuc  order by left(TimeSave,11) asc",conn)
                    plt.rc('font', size=14)
                    pivot_df = df7.pivot_table(index='Time', columns=['KhuVuc'], values='Count', aggfunc='sum', fill_value=0)
                    ax = pivot_df.plot(kind='bar', figsize=(20, 10), width=0.8,color=quality_color_mapping)
                    # ax.legend(bbox_to_anchor =(1, 1.09), ncol = 2)
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    ax.legend(title= 'Quality',  bbox_to_anchor= (1, 1.24), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title("Manual Scan Data")
                    plt.savefig('static/images/Manualscan.png')
                
                return render_template('pages/Manualscan.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend,areascan=areascan)
            else:
                return render_template('/pages/Manualscan.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)  
    else:
        return redirect(url_for('login'))


@app.route('/download/Manualscan',methods=['POST'])
def download_Manualscan():
    # global areascan
    # payload = request.get_json()
    # areascan = request.form['areapos']
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
   
    areascan=request.form['areapos']
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    
    print(areascan,timestart,timeend)
    
    if areascan =='All':
        df8 = pd.read_sql("select * FROM [QC].[dbo].[Scan_Repair_Data] where TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"' order by TimeSave asc", conn)
        print("select * FROM [QC].[dbo].[Scan_Repair_Data] where TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"' order by TimeSave asc")
    else:
        df8 = pd.read_sql("select * FROM [QC].[dbo].[Scan_Repair_Data] where KhuVuc like N'%"+areascan+"%'and  TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"' order by TimeSave asc", conn)
        print("select * FROM [QC].[dbo].[Scan_Repair_Data] where KhuVuc like N'%"+areascan+"%'and  TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"' order by TimeSave asc")
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df8.to_excel(writer, sheet_name='Manualscan',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Manualscan_Data.xls"})
@app.route("/Welding", methods=['POST', 'GET'])
def open_Data():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Welding"
        if check_secr(14): 
            if request.method == 'POST':
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                cursor = conn.cursor()
                # data = laser_all_data("","","",timestart,timeend)
                cursor.execute("SELECT Top(15000)* From [QC].[dbo].[Laser_Welding] where Time_Finish >'"+timestart+"' and Time_Start <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    quality_color_mapping = {
                        ('Gá Tổng Hợp'): 'grey',
                        ('Kích Thước Khác'): 'darkgreen',
                        ('Ngoại Quan'): 'darkblue',
                        ('Pin Ren'): 'gold',
                        ('Ren Tay'): 'orange',
                        
                    }
                    # if areascan == 'All':
                    #     df7 = pd.read_sql("SELECT left(TimeSave,11) as 'Time',KhuVuc,  COUNT(DMC) AS 'Count'FROM [QC].[dbo].[Scan_Repair_Data] where TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"'  GROUP BY left(TimeSave,11),KhuVuc  order by left(TimeSave,11) asc",conn)
                    # else:
                    #     df7 = pd.read_sql("SELECT left(TimeSave,11) as 'Time',KhuVuc,  COUNT(DMC) AS 'Count'FROM [QC].[dbo].[Scan_Repair_Data] where KhuVuc like N'%"+areascan+"%' and TimeSave >='"+timestart+"' and TimeSave <='"+timeend+"'  GROUP BY left(TimeSave,11),KhuVuc  order by left(TimeSave,11) asc",conn)
                    df7 = pd.read_sql("SELECT left(Time_Finish,11) as 'Time',  COUNT(DMC_OC) AS 'Count'FROM  [QC].[dbo].[Laser_Welding] where  Time_Finish >='"+timestart+"' and Time_Finish <='"+timeend+"'  GROUP BY left(Time_Finish,11)  order by left(Time_Finish,11) asc",conn)

                    plt.rc('font', size=14)
                    pivot_df = df7.pivot_table(index='Time', values='Count', aggfunc='sum', fill_value=0)
                    ax = pivot_df.plot(kind='bar', figsize=(20, 10), width=0.8,color='green')
                    # ax.legend(bbox_to_anchor =(1, 1.09), ncol = 2)
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    ax.legend(title= 'Quality',  bbox_to_anchor= (1, 1.24), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title("Welding Data")
                    plt.savefig('static/images/Welding.png')
                return render_template('pages/Moi.html',user=user,process=process,post_flag = True,data = data,soluong = soluong, timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Moi.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/ThreadVerification", methods=['POST', 'GET'])
def open_ThreadVerification():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="ThreadVerification"
        if check_secr(20): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)* FROM [QC].[dbo].[ThreadVerification] where Timestart >'"+timestart+"' and Timestart <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    quality_color_mapping = {
                        ('Machine1', 'NG'): 'darkred',
                        ('Machine1', 'OK'): 'darkgreen',
                        ('Machine1', 'Special'): 'darkblue',
                        ('Machine1', 'Return'): 'gold',
                        ('Machine2', 'NG'): 'red',
                        ('Machine2', 'OK'): 'lime',
                        ('Machine2', 'Special'): 'blue',
                        ('Machine2', 'Return'): 'yellow'
                    }
                    df7 = pd.read_sql(
                    "   SELECT left(Timestart,11) as 'Time',Machine, Quality, COUNT(Quality) AS 'Count'FROM [QC].[dbo].[ThreadVerification] where Timestart >'"+timestart+"' and Timestart <'"+timeend+"' and Quality != '' GROUP BY left(Timestart,11),Quality,Machine  order by left(Timestart,11) desc",
                    conn)
                    plt.rc('font', size=14)
                    pivot_df = df7.pivot_table(index='Time', columns=['Machine', 'Quality'], values='Count', aggfunc='sum', fill_value=0)
                    ax = pivot_df.plot(kind='bar', figsize=(20, 10), width=0.8,color=quality_color_mapping)
                    # ax.legend(bbox_to_anchor =(1, 1.09), ncol = 2)
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    ax.legend(title= 'Quality',  bbox_to_anchor= (1, 1.24), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title("TOTALS OF NG & OK PRODUCTS  ")
                    plt.savefig('static/images/ThreadVerification.png')
                return render_template('pages/ThreadVerification.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/ThreadVerification.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/ThreadVerification/realtime", methods=['POST', 'GET'])
def open_ThreadVerificationrealtime():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="ThreadVerification"
        if check_secr(20): 
            return render_template('/pages/ThreadVerificationrealtime.html',user=user,process=process)
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/ThreadVerification/rt2D/<machine>", methods=['GET'])
def open_ThreadVerificationrt2D(machine):
    if request.cookies.get("login_status")=='ok':
        with open("static/airleakage/pointmap.json", 'r') as json_file:
            points = json.load(json_file)
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; Database=QC; UID=sa; PWD=1234; Trusted_Connection=No;')
        data = pd.read_sql("select * from [2D_Thread] where machineno = '"+machine+"'",conn)
        mahang = data.loc[0,"Product_type"]  
        img = cv2.imread('static/airleakage/'+mahang+'.png')
        # len(points[mahang])
        
        dt = pd.read_sql("select top(1) Quality from [QC].[dbo].[ThreadVerification] where [DMC_product] = '"+data.loc[0,"DMC"]+"' order by Timestart desc",conn)
        if len(dt)>0:
            if dt.loc[0,"Quality"] == "OK":
                cv2.rectangle(img,(int(img.shape[1]/2)-210,70),(int(img.shape[1]/2)+390,110), (0, 255, 0), -1)
            else:
                cv2.rectangle(img,(int(img.shape[1]/2)-210,70),(int(img.shape[1]/2)+390,110), (0, 0, 255), -1)
        else:
            cv2.rectangle(img,(int(img.shape[1]/2)-210,70),(int(img.shape[1]/2)+390,110), (200, 200, 200), -1)
        cv2.putText(img,data.loc[0,"DMC"],(int(img.shape[1]/2)-200,100),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2,cv2.LINE_AA)  
        for i in range(int(data.loc[0,"Hole_no"])):
            point = points[mahang][i]   
                
            if int(data.loc[0,"Hole_stt"+str(i+1)]) == 1:
                color = (0,255,0)
            else:
                color = (0,0,255)  
            cv2.circle(img,(point["position"][0],point["position"][1]),point["radius"],color,3)
            if mahang =='A2303121':
                if int(data.loc[0,"Hole_no"]) < len(points[mahang]):
                    point2 = points[mahang][int(data.loc[0,"Hole_no"])] 
                    if datetime.datetime.now().second % 2 == 0:
                        cv2.circle(img,(point2["position"][0],point2["position"][1]),point["radius"],(0,255,255),3)
            else:        
                if int(data.loc[0,"Hole_no"]) < len(points[mahang]):
                    point2 = points[mahang][int(data.loc[0,"Hole_no"])] 
                    if datetime.datetime.now().second % 2 == 0:
                        cv2.circle(img,(point2["position"][0],point2["position"][1]),point["radius"],(0,255,255),3)
                # else:
                #     cv2.circle(img,(point2["position"][0],point2["position"][1]),point["radius"],(0,0,255),3)
            cv2.putText(img,str(round(float(data.loc[0,"Hole_tor"+str(i+1)]),4)),(point["position"][0]-50,point["position"][1]-20),cv2.FONT_HERSHEY_SIMPLEX,0.5,(0,0,0),5,cv2.LINE_AA) 
            cv2.putText(img,str(round(float(data.loc[0,"Hole_tor"+str(i+1)]),4)),(point["position"][0]-50,point["position"][1]-20),cv2.FONT_HERSHEY_SIMPLEX,0.5,color,2,cv2.LINE_AA)
        filename = 'static/airleakage/'+machine+str(data.loc[0,"Hole_no"])+'.png'
        cv2.imwrite(filename,img)
        return send_file(filename, mimetype='image/png')
    else:
        return redirect(url_for('login'))
@app.route("/ThreadVerification/getinfo/<dmc>", methods=['GET'])
def open_ThreadVerificationgetinfo(dmc):
    if request.cookies.get("login_status")=='ok':
        with open("static/airleakage/pointmap.json", 'r') as json_file:
            points = json.load(json_file)
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; Database=QC; UID=sa; PWD=1234; Trusted_Connection=No;')
        data = pd.read_sql("select top(1) * from [QC].[dbo].[ThreadVerification] where id = '"+dmc+"' order by id desc",conn)
        mahang = data.loc[0,"ProductName"]  
        img = cv2.imread('static/airleakage/'+mahang+'.png')
        
        
        if data.loc[0,"Quality"] == "OK":
            cv2.rectangle(img,(int(img.shape[1]/2)-210,70),(int(img.shape[1]/2)+390,110), (0, 255, 0), -1)
        else:
            cv2.rectangle(img,(int(img.shape[1]/2)-210,70),(int(img.shape[1]/2)+390,110), (0, 0, 255), -1)
        cv2.putText(img,data.loc[0,"DMC_product"],(int(img.shape[1]/2)-200,100),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2,cv2.LINE_AA) 
        errorhole = []
        if len(data.loc[0,"Status"])> 0:
            err = str(data.loc[0,"Status"]).split("HOLE_")
            for hole in err:
                try:
                    errorhole.append(int(hole.strip()))
                except:
                    print("error",hole)
        if mahang == 'A2303121':
            for i in range(7):
                point = points[mahang][i]   
                
                if (i+1) in errorhole:
                    color = (0,0,255)
                else:
                    color = (0,255,0)  
                cv2.circle(img,(point["position"][0],point["position"][1]),point["radius"],color,3)
                cv2.putText(img,str(round(float(data.loc[0,"Torque_H"+str(i+1)]),4)),(point["position"][0]-50,point["position"][1]-20),cv2.FONT_HERSHEY_SIMPLEX,0.5,(0,0,0),5,cv2.LINE_AA)
                cv2.putText(img,str(round(float(data.loc[0,"Torque_H"+str(i+1)]),4)),(point["position"][0]-50,point["position"][1]-20),cv2.FONT_HERSHEY_SIMPLEX,0.5,color,2,cv2.LINE_AA)  
        else:
            for i in range(11):
                point = points[mahang][i]   
                
                if (i+1) in errorhole:
                    color = (0,0,255)
                else:
                    color = (0,255,0)  
                cv2.circle(img,(point["position"][0],point["position"][1]),point["radius"],color,3)
                cv2.putText(img,str(round(float(data.loc[0,"Torque_H"+str(i+1)]),4)),(point["position"][0]-50,point["position"][1]-20),cv2.FONT_HERSHEY_SIMPLEX,0.5,(0,0,0),5,cv2.LINE_AA)
                cv2.putText(img,str(round(float(data.loc[0,"Torque_H"+str(i+1)]),4)),(point["position"][0]-50,point["position"][1]-20),cv2.FONT_HERSHEY_SIMPLEX,0.5,color,2,cv2.LINE_AA)  
        filename = 'static/airleakage/showpic.png'
        cv2.imwrite(filename,img)
        return send_file(filename, mimetype='image/png')
    else:
        return redirect(url_for('login'))
@app.route("/ThreadVerification/top10/<machine>", methods=['GET'])
def open_ThreadVerificationtop10(machine):
    if request.cookies.get("login_status")=='ok':
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
        data = pd.read_sql("select top(10) * from [QC].[dbo].[ThreadVerification] where [Machine] = '"+machine+"' order by Timestart desc",conn)
        edata = {}
        # edata["Product"]=[]
        for i in range(len(data)):
            edata["data"+str(data.loc[i,'id'])]={"id":str(data.loc[i,'id']),
                                          "OP_name":data.loc[i,'OP_name'],
                                          "Machine":data.loc[i,'Machine'],
                                          "ProductName":data.loc[i,'ProductName'],
                                          "OP_name":data.loc[i,'OP_name'],
                                          "DMC_product":data.loc[i,'DMC_product'],
                                          "Timestart":data.loc[i,'Timestart'],
                                          "Timefinish":data.loc[i,'Timefinish'],
                                          "Quality":data.loc[i,'Quality'],
                                          "Status":data.loc[i,'Status'],
                                          "Note_life_time":data.loc[i,'Note_life_time']
                                          }
        # data = data.to_json()
        return jsonify(edata)
    else:
        return redirect(url_for('login'))
@app.route("/ThreadVerification_update", methods=['POST'])
def ThreadVerification_update():
    if check_secr(21): 
        username = request.cookies.get("user_name")
        data = request.get_json()
        password = data.get("pass")
        
        dmc = data.get("dmc")
        result = data.get("result")
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=Auto; Trusted_Connection=No;',timeout=1)
        confirm_user = pd.read_sql("SELECT Top(1) * FROM [Auto].[dbo].[Web_User_Test] where username = '"+username+"' and password = '"+password+"'" , conn)
        print(confirm_user)
        if len(confirm_user) == 1:
            cursor = conn.cursor()
            cursor.execute("Update [QC].[dbo].[ThreadVerification] set Quality ='"+result+"' where DMC_Product ='"+dmc+"' ")
            cursor.commit()
            return "OK"
        else:
            return  'Wrong Password'
    else:
        return "You don't have permission"
@app.route("/ShotBlasting", methods=['POST', 'GET'])
def open_ShotBlasting():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="ShotBlasting"
        if check_secr(13): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=Auto; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)* FROM [Auto].[dbo].[ShotBlasting] where Time_Take_out >'"+timestart+"' and Time_Take_out <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df8 = pd.read_sql(
                    "     SELECT left(Time_Take_out,11) as 'Time', MachineNo, COUNT(MachineNo) AS 'Count'FROM [Auto].[dbo].[ShotBlasting] where Time_Take_out >'"+timestart+"' and Time_Take_out <'"+timeend+"' GROUP BY left(Time_Take_out,11),MachineNo  order by left(Time_Take_out,11) desc",
                    conn)
                    plt.rc('font', size=14)
                    
                    df8 = df8.pivot_table(index='Time', columns='MachineNo')
                    ax = df8.plot(kind='bar', y='Count', figsize=(20, 10),color=['navy', 'darkgreen', 'red', 'gold'], width = 0.7)
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    ax.legend(title= 'MachineNo',  bbox_to_anchor= (1, 1.17), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title(" THE NUMBER OF PRODUCTS PER MACHINE ")
                    plt.savefig('static/images/ShotBlasting.png')
                return render_template('pages/ShotBlasting.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/ShotBlasting.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))

@app.route("/Cutting", methods=['POST', 'GET'])
def open_Cutting():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Cutting"
        if check_secr(13): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=HXL; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)* FROM [HXL].[dbo].[DSA_cutting] where [Time] >'"+timestart+"' and [Time] <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df8 = pd.read_sql(
                    "     SELECT left([Time],11) as 'Time', [cutting_name], COUNT([cutting_name]) AS 'Count'FROM [HXL].[dbo].[DSA_cutting] where [Time] >'"+timestart+"' and [Time] <'"+timeend+"' GROUP BY left([Time],11),[cutting_name]  order by left([Time],11) desc",
                    conn)
                    plt.rc('font', size=14)
                    
                    df8 = df8.pivot_table(index='Time', columns='cutting_name')
                    ax = df8.plot(kind='bar', y='Count', figsize=(20, 10),color=['navy', 'darkgreen', 'red', 'gold', 'yellow'], width = 0.7)
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    ax.legend(title= 'MachineNo',  bbox_to_anchor= (1, 1.17), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title(" THE NUMBER OF PRODUCTS CUTTING MACHINE ")
                    plt.savefig('static/images/Cutting.png')
                return render_template('pages/Cutting.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/Cutting.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route('/download/Cutting',methods=['POST'])
def download_Cutting():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=HXL; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df8 = pd.read_sql("SELECT * FROM [HXL].[dbo].[DSA_cutting] where [Time] >'"+timestart+"' and [Time] <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df8.to_excel(writer, sheet_name='Cutting',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Cutting_Data.xls"})
@app.route("/AirguageGC", methods=['POST', 'GET'])
def open_AirguageGC():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="AirguageGC"
        if check_secr(18): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=GC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)* FROM [GC].[dbo].[Measure_Diameter] where Time_finish >'"+timestart+"' and Time_finish <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df7 = pd.read_sql(
                    "   SELECT left(Time_finish,11) as 'Time', Result, COUNT(Result) AS 'Count'FROM [GC].[dbo].[Measure_Diameter] where Time_finish >'"+timestart+"' and Time_finish <'"+timeend+"' GROUP BY left(Time_finish,11), Result  order by left(Time_finish,11) desc",
                    conn)
                    plt.rc('font', size=14)
                    
                    df7 = df7.pivot_table(index='Time', columns='Result')
                    ax = df7.plot(kind='bar', y='Count', figsize=(20, 10), color=['red', 'darkgreen'], width = 0.7)
                    # ax.legend(bbox_to_anchor =(1, 1.09), ncol = 2)
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    ax.legend(title= 'Result',  bbox_to_anchor= (1, 1.12), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title("TOTALS OF NG & OK PRODUCTS  ")
                    plt.savefig('static/images/MeasureDiameterGC.png')
                return render_template('pages/MeasureDiameterGC.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/MeasureDiameterGC.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))    
@app.route("/AirguageQC", methods=['POST', 'GET'])
def open_AirguageQC():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="AirguageQC"
        if check_secr(19): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)* FROM [QC].[dbo].[Measure_Diameter] where Time_finish >'"+timestart+"' and Time_finish <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df7 = pd.read_sql(
                    "   SELECT left(Time_finish,11) as 'Time', Result, COUNT(Result) AS 'Count'FROM [QC].[dbo].[Measure_Diameter] where Time_finish >'"+timestart+"' and Time_finish <'"+timeend+"' GROUP BY left(Time_finish,11), Result  order by left(Time_finish,11) desc",
                    conn)
                    plt.rc('font', size=14)
                    
                    df7 = df7.pivot_table(index='Time', columns='Result')
                    ax = df7.plot(kind='bar', y='Count', figsize=(20, 10), color=['red', 'darkgreen', 'orange','blue'], width = 0.7)
                    # ax.legend(bbox_to_anchor =(1, 1.09), ncol = 2)
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    ax.legend(title= 'Result',  bbox_to_anchor= (1, 1.12), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title("TOTALS OF NG & OK PRODUCTS  ")
                    plt.savefig('static/images/MeasureDiameterQC.png')
                return render_template('pages/MeasureDiameterQC.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/MeasureDiameterQC.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))  
    


@app.route("/ScanQR", methods=['POST', 'GET'])
def open_ScanQR():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="ScanQR"
        if check_secr(15): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)* FROM [QC].[dbo].[ScanQR] where Time_scan_tray >'"+timestart+"' and Time_scan_tray <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df7 = pd.read_sql(
                    "select left(Time_scan_tray,11) as 'Time', Compare, Count(compare) as 'Count' from [QC].[dbo].[ScanQR] where Time_scan_tray >'"+timestart+"' and Time_scan_tray <'"+timeend+"' GROUP BY left(Time_scan_tray,11), Compare  order by left(Time_scan_tray,11) desc",
                    conn)
                    plt.rc('font', size=14)
                    
                    df7 = df7.pivot_table(index='Time', columns='Compare')
                    ax = df7.plot(kind='bar', y='Count', figsize=(20, 10), color=['red','darkgreen'], width = 0.5)
                    # ax.legend(bbox_to_anchor =(1, 1.09), ncol = 2)
                    for container in ax.containers:
                        ax.bar_label(container,rotation=0)
                    #ax.legend(title= 'Quality',  bbox_to_anchor= (1, 1.12), ncol = 2)
                    plt.xticks(rotation=45)
                    plt.subplots_adjust(bottom=0.213, right=0.996, left=0.032)
                    plt.title("TOTALS PRODUCTS  ")
                    plt.savefig('static/images/ScanQR.png')
                return render_template('pages/ScanQR.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/ScanQR.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/TiltMeasurement", methods=['POST', 'GET'])
def open_TiltMeasurement():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="TiltMeasurement"
        if check_secr(16): 
            if request.method == 'POST':
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                
                cursor = conn.cursor()
                cursor.execute("SELECT Top(15000)* FROM [QC].[dbo].[TiltMeasurement] where Time_finish >'"+timestart+"' and Time_finish <'"+timeend+"'")
                data =cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df1 = pd.read_sql(
                        "   SELECT left(Time_finish,11)as Time,Result, COUNT(Result) AS 'Count' FROM [QC].[dbo].[TiltMeasurement] where  Time_finish >'"+timestart+"' and Time_finish < '"+timeend+"' GROUP BY left(Time_finish,11),Result order by left(Time_finish,11)",
                        conn)
                    plt.rc('font', size=14)
                    
                    df1 = df1.pivot_table(index='Time', columns='Result')
                    
                    ax1 = df1.plot(kind='bar', y='Count', figsize=(20, 10), color=['red', 'darkgreen'], width = 0.7)
                    for container in ax1.containers:
                        ax1.bar_label(container, rotation=0)
                    ax1.legend(title= 'Result',  bbox_to_anchor= (1, 1.13), ncol = 2)
                    plt.subplots_adjust(bottom=0.265, right=0.975, left=0.078)
                    # plt.figure(figsize=(2, 2))
                    plt.xticks(rotation=45)
                    plt.title("A2012003/004-Tilt Measurement")
                    plt.savefig('static/images/TiltMeasurement.png')
                return render_template('pages/TiltMeasurement.html',user=user,process=process,post_flag = True,data = data,page=1,soluong=soluong,timestart=timestart,timeend=timeend)
            else:
                return render_template('/pages/TiltMeasurement.html',user=user,process=process,timestart="",timeend="")
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route('/download/ShotBlasting',methods=['POST'])
def download_ShotBlasting():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=Auto; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df8 = pd.read_sql("SELECT*From[Auto].[dbo].[ShotBlasting] where Time_Take_out >'"+timestart+"' and Time_Take_out <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df8.to_excel(writer, sheet_name='ShotBlasting',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=ShotBlasting_Data.xls"})
@app.route('/download/OpticalDivision',methods=['POST'])
def download_OpticalDivision():

    wb_sample = load_workbook('static/samplespc.xlsx')
    sheet=wb_sample.active
    
    filegroup = request.cookies.get("grpname")   
    filename = request.cookies.get('filename') 
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=SPC; Trusted_Connection=No;',timeout=1)
    datafileid = session.get('datafileid')
    datafileid_str = str(datafileid["data"][0])
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    print('fileid',datafileid_str)

    conn1 = pyodbc.connect('Driver={SQL Server}; Server=192.168.9.246; uid=sa; pwd=saa; Database=SPC; Trusted_Connection=No;', timeout=1)
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234; Database=SPC; Trusted_Connection=No;', timeout=1)
    if filegroup =='DS50000-Heat-Treatment':
        data1 = pd.read_sql("SELECT * FROM GetControltorque() where fileid = '"+datafileid_str+"'",conn1)
        data2 = pd.read_sql("select * from spc.dbo.getcontrol821_torque('"+datafileid_str+"', '"+timestart+"', '"+timeend+"')",conn)
        data3 = pd.read_sql("SELECT * FROM spctorque() where fileid = '"+datafileid_str+"' and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'",conn)
        wb_sample = load_workbook('static/sampletorque.xlsx')
        # post = 0
    elif filegroup =='DC10200-Casting':

        data1 = pd.read_sql("SELECT * FROM GetControl3() where fileid = '"+datafileid_str+"'",conn1)
        data2 = pd.read_sql("select * from spc.dbo.getcontrol821_v3('"+datafileid_str+"', '"+timestart+"', '"+timeend+"')",conn)
        data3 = pd.read_sql("SELECT * FROM spc6() where fileid = '"+datafileid_str+"' and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'",conn)
        wb_sample = load_workbook('static/samplespc.xlsx')
    sheet=wb_sample.active
    rs = pd.concat([data1, data2], ignore_index=True)
    newdata = {}
    for col in rs.columns[2:]:
        newdata[col] = []
        splqlt = len(data3) -1
        test = 0
        sps = int(rs.loc[4,col])
        if sps == 1:
            try:
            
                for i in range(splqlt):
                    test += abs(data3.loc[i+1,col] - data3.loc[i,col])
                dev = test/splqlt/1.128
                
                newdata[col].append(round(dev,4))
                newdata[col].append(int(splqlt+1))
            except:
                
                newdata[col].append('--')
                newdata[col].append(0)
        else:
            try:
                k107 = 1
                for i in range(0,splqlt+1,sps):
                    savg = np.average(data3.loc[i:i+sps-1,"C"])
                    sumd = 0
                    for j in data3.loc[i:i+sps-1,"C"]:
                        sumd += (float(j)-savg)**2
                    # print(sumd)
                    test += sumd
                    k107 += len(data3.loc[i:i+sps-1,"C"]) - 1
                    # print(k107) 
                l108 = math.sqrt(test/(k107-1))
                k109 = calculate_standard_deviation(k107)
                dev = l108/k109
                newdata[col].append(round(dev,4))
                newdata[col].append(int(splqlt+1))
            except:
                newdata[col].append('--')
                newdata[col].append(0)
        #Ca
        try:        
            newdata[col].append(round((2*float(rs.loc[5,col]) - float(rs.loc[1,col]) - float(rs.loc[2,col]))/(float(rs.loc[1,col]) - float(rs.loc[2,col])),4)) #ca=2avg-usl-lsl/usl-lsl
        except:
            newdata[col].append('--')
        #Cp
        
        try:        
            newdata[col].append(round((float(rs.loc[1,col]) - float(rs.loc[2,col]))/(6*dev),4)) #cp=(usl-lsl)/6*dev   USL, LSL != '--'
        except:
            try:        
                newdata[col].append(round(abs((float(rs.loc[5,col]) - float(rs.loc[1,col])))/(3*dev),4)) # cp = |avg-usl|/3.dev USL != '--'
            except:
                try:
                    newdata[col].append(round(abs((float(rs.loc[5,col]) - float(rs.loc[2,col])))/(3*dev),4)) # cp = |avg-lsl|/3.dev LSL != '--'
                except:
                    newdata[col].append('--') #LSL = USL = '--'

        try:        
            newdata[col].append(round(((float(rs.loc[1,col]) - float(rs.loc[2,col]))/(6*dev))*(1-abs((2*float(rs.loc[5,col]) - float(rs.loc[1,col]) - float(rs.loc[2,col]))/(float(rs.loc[1,col]) - float(rs.loc[2,col])))),4)) #Cpk = ((usl-lsl)/6*dev) * (1-|2avg-usl-lsl/usl-lsl|)
        except:
            try:        
                newdata[col].append(round(abs((float(rs.loc[5,col]) - float(rs.loc[1,col])))/(3*dev),4)) # cpk = |avg-usl|/3.dev USL != '--'
            except:
                try:
                    newdata[col].append(round(abs((float(rs.loc[5,col]) - float(rs.loc[2,col])))/(3*dev),4)) # cpk = |avg-lsl|/3.dev LSL != '--'
                except:
                    newdata[col].append('--') #LSL = USL = '--'
        # P
        count = 0
        for x in range(splqlt):
            try:
                if float(data3.loc[x,col]) > float(data1.loc[1,col]):
                    print(col, data3.loc[x,col])
                    count += 1
                    for coll in range(len(data3.columns)):
                        if data3.columns[coll] == col:
                            sheet.cell(28+x,coll-1).font = Font(color = "FF0000")
            except:
                pass
            try:
                if float(data3.loc[x,col]) < float(data1.loc[2,col]):
                    print(col, data3.loc[x,col])
                    count += 1
                    for coll in range(len(data3.columns)):
                        if data3.columns[coll] == col:
                            sheet.cell(28+x,coll-1).font = Font(color = "FF0000")
            except:
                pass
        print(col, count)
        if len(data3) >0:
            newdata[col].append(str(round(100*count/len(data3),2))+'%') 
        else:
            newdata[col].append(0)

    newdata['Control_Item'] = ['Std.Dev','Qty','Ca','Cp','Cpk','P']
    newdata['FileID'] = [datafileid_str]*6
    print(newdata)
    dt1 = pd.DataFrame(newdata)
    rs2_list = pd.concat([rs, dt1], ignore_index=True)

    df7 = rs2_list
    columns_to_drop = ['FileID','Control_Item']
    df7 = df7.drop(columns_to_drop, axis=1)

    if filegroup =='DS50000-Heat-Treatment':
        df8 = pd.read_sql("select * from SPCtorque() where FileID = '"+datafileid_str+"' and VCTRLID in (SELECT [VCTRLID] FROM [SPC].[dbo].[HeatTreament_Info]) and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'", conn)
    elif filegroup =='DC10200-Casting':
    # df8 = pd.read_sql("select Datetimecreate, [C], [Si], [Mn], [P], [S], [Ni], [Cr], [Mo], [Cu], [Ti], [V], [Pb], [W], [Al], [Co], [Nb], [As], [Sn], [Sb], [B], [Bi], [Ca], [Zn], [N], [Ce], [Mg], [Ta], [Zr], [Ti+Nb], [Fe%], [CEF], [Other], [A1_Appearance] from spc6() where FileID = '"+datafileid_str+"' and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'", conn)
        df8 = pd.read_sql("select * from spc6() where FileID = '"+datafileid_str+"' and DateTimeCreate >'"+timestart+"' and DateTimeCreate <'"+timeend+"'", conn)
    columns_to_drop = ['FileID', 'VCTRLID']
    df8 = df8.drop(columns_to_drop, axis=1)
    my_red = openpyxl.styles.colors.Color(rgb='00CCCCFF')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    border1 = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
    border2 = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thin'))
    border3 = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thick'))
    # border4 = Border(left=Side(style='thin'))

    for row in range(len(df7)):         
        for col in range(len(df7.columns)):
            sheet.cell(7+row,col+2).value = df7.loc[row,df7.columns[col]]
    for row in range(len(df8)):
        for col in range(len(df8.columns)):
            # sheet.cell(28+row,1).border =border4
            sheet.cell(1,2).value = filegroup
            sheet.cell(2,2).value = filename
            sheet.cell(3,2).value = df8.loc[0,'Material']
            sheet.cell(5,2).value = str(df8.loc[0,'DateTimeCreate'])[:-7] +' ~ '+ str(df8.loc[len(df8)-1,'DateTimeCreate'])[:-7]
            sheet.cell(28+row,col+1).value = df8.loc[row,df8.columns[col]]
            sheet.cell(28+row,1).fill = my_fill
            sheet.cell(28+row,col+1).border = border
            sheet.cell(27+len(df8),col+1).border = border1
            sheet.cell(28+row,len(df8.columns)).border = border2

    sheet.cell(27+len(df8),len(df8.columns)).border = border3        
#CCCCFF
    # Bỏ hai cột khỏi dataframe
    
    output = io.BytesIO()
    # # writer = pd.ExcelWriter(output, engine='xlwt')  # add a sheet
    # writer = pd.ExcelWriter(output, engine='xlsxwriter')    # add a sheet
    # # wb_sample = pd.ExcelWriter(output, engine='xlsxwriter')
    # # df8.to_excel(writer, sheet_name='OpticalDivision',index=False)
    # # df7.to_excel(writer, sheet_name='OpticalDivision', startrow=0, index=False)
    # # df8.to_excel(writer, sheet_name='OpticalDivision', startrow=df7.shape[0]+2, index=False)
    wb_sample.save(output)
    # writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=OpticalDivisionData_"+datetime.datetime.now().strftime('%Y%m%d_%Hh%Mp')+".xls"})

@app.route("/TempHumid", methods=['POST', 'GET'])
def open_Temphumid():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process = "TempHumid"
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)

        # cursor = conn.cursor()
        # cursor.execute("SELECT Top(15000)*  From [QC].[dbo].[DONG_BIN] where TimeFinish >'"+timestart+"' and TimeFinish <'"+timeend+"'")
        # #print("SELECT Top(15000)* From[QC].[dbo].[air_tight] where Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'")
        # data =cursor.fetchall()
        data_areaname= pd.read_sql("select distinct(area_name) from [QC].[dbo].[temp_humid_setting]",conn)
        count_areaname= pd.read_sql("select count(distinct(area_name)) as count from [QC].[dbo].[temp_humid_setting]",conn)
        data_areaname1=data_areaname['area_name'].tolist()
        count_areaname1=count_areaname.loc[0,'count']
        print(count_areaname1)
        if check_secr(31): 
            if request.method == 'POST':
                # conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                areaname = request.form['coatingpos']
                print(areaname)
                cursor = conn.cursor()
                cursor.execute("select top 10000 Area_name, Temp, Humid, Time_get from [Temp_Humid_Factory] left join [temp_humid_setting] on [Temp_Humid_Factory].Area = [temp_humid_setting].Area where Area_name = N'"+areaname+"' and Time_get >='"+timestart+"' and Time_get <='"+timeend+"'")
                # #print("SELECT Top(15000)* From[QC].[dbo].[air_tight] where Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'")
                data = cursor.fetchall()
                soluong = len(data)
                if soluong >0:
                    df1 = pd.read_sql("select Area_name, Temp, Humid, Time_get from [Temp_Humid_Factory] left join [temp_humid_setting] on [Temp_Humid_Factory].Area = [temp_humid_setting].Area where Area_name = N'"+areaname+"' and Time_get >='"+timestart+"' and Time_get <='"+timeend+"' order by Time_get desc",conn)
                    df1["Time_get"] = pd.to_datetime(df1["Time_get"], format='%Y-%m-%d %H:%M:%S.%f')
                    df1 = df1.set_index('Time_get')
                    target = pd.read_sql("select * from [QC].[dbo].[temp_humid_setting] where Area_name = N'"+areaname+"'",conn)
                    temp_min = target.loc[0,'temp_min']  # Thay thế bằng giá trị nhiệt độ mục tiêu của bạn
                    temp_max = target.loc[0,'temp_max']  # Thay thế bằng giá trị độ ẩm mục tiêu của bạn
                    humid_min= target.loc[0,'humid_min']
                    humid_max= target.loc[0,'humid_max']
                    plt.figure(figsize=(10, 6))
                    plt.scatter(df1.index, df1['Temp'], label='Temperature',color='orange',  s=5)
                    plt.scatter(df1.index, df1['Humid'], label='Humidity',color='blue',  s=5)

                    plt.axhline(y=temp_min, color='red', linestyle='-', label='Temp Min')
                    plt.axhline(y=temp_max, color='green', linestyle='-', label='Temp Max')
                    plt.axhline(y=humid_min, color='yellow', linestyle='-', label='Humidity Min')
                    plt.axhline(y=humid_max, color='purple', linestyle='-', label='Humidity Max')

                    plt.text(df1.index[0], temp_min, f'          {int(temp_min)}', color='r', va='bottom', ha='left')
                    plt.text(df1.index[0], temp_max, f'          {int(temp_max)}', color='g', va='bottom', ha='left')
                    plt.text(df1.index[0], humid_min, f'          {int(humid_min)}', color='yellow', va='bottom', ha='left')
                    plt.text(df1.index[0], humid_max, f'          {int(humid_max)}', color='purple', va='bottom', ha='left')


                    plt.xlabel('Datetime')
                    plt.ylabel('Values')
                    plt.title(areaname+' Temperature & Humidity',loc='left')
                    plt.legend(bbox_to_anchor= (1, 1.1), ncol = 3,fontsize= 6)
                    plt.xticks(rotation=45, fontsize= 8)
                    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d %H:%M'))  
                    plt.tight_layout()
                    plt.savefig('static/images/Temphumid.png')
                
                return render_template('pages/Temphumid.html',user=user,process=process,post_flag = True,soluong=soluong,data = data,page=1,timestart=timestart,timeend=timeend,data_areaname1=data_areaname1,count_areaname1=count_areaname1,areaname=areaname)
            else:
                return render_template('/pages/Temphumid.html',user=user,process=process,timestart="",timeend="",data_areaname1=data_areaname1,count_areaname1=count_areaname1)
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))

@app.route('/download/TempHumid',methods=['POST'])
def download_TempHumid():
    # global areaname
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    areaname = request.form['areaname']
    # print(timestart)
    df7 = pd.read_sql("select top 10000 Area_name, Temp, Humid, Time_get from [Temp_Humid_Factory] left join [temp_humid_setting] on [Temp_Humid_Factory].Area = [temp_humid_setting].Area where Area_name = N'"+areaname+"' and Time_get >='"+timestart+"' and Time_get <='"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df7.to_excel(writer, sheet_name='TempHumid',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": f"attachment;filename=TempHumid_Data_{areaname}.xls"})

@app.route("/TempHumidRealtime", methods=['POST', 'GET'])
def open_TempHumidRealtime():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(32): 
            process="TempHumidRealtime"
            conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
            
            cursor = conn.cursor()
            cursor.execute("select Area_name, Temp, Humid, Temp_Min, Temp_Max, Humid_Min, Humid_Max from [temp_humid_realtime] left join [temp_humid_setting] on [temp_humid_realtime].Area = [temp_humid_setting].Area ")
            data =cursor.fetchall()
            print(data)
            return render_template('/pages/TempHumidRealtime.html',user=user,process=process,tempdata = data)
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))

@app.route("/TempHumidRealtime/update", methods=['POST', 'GET'])
def update_TempHumidRealtime():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    
    data = pd.read_sql("select Area_name, Temp, Humid, Temp_Min, Temp_Max, Humid_Min, Humid_Max from [temp_humid_realtime] left join [temp_humid_setting] on [temp_humid_realtime].Area = [temp_humid_setting].Area ",conn)
    edata = {}
    for i in range(len(data)):
        edata[data.loc[i,"Area_name"]]={
            "Temp":round(data.loc[i,"Temp"],1),
            "Humid":round(data.loc[i,"Humid"],1),
            "Temp_Min":round(data.loc[i,"Temp_Min"],1),
            "Temp_Max":round(data.loc[i,"Temp_Max"],1),
            "Humid_Min":round(data.loc[i,"Humid_Min"],1),
            "Humid_Max":round(data.loc[i,"Humid_Max"],1)
        }
    # data = data.to_json()
    return jsonify(edata)



@app.route('/download/ThreadVerification',methods=['POST'])
def download_ThreadVerification():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    print(timestart)
    df7 = pd.read_sql("SELECT*From[QC].[dbo].[ThreadVerification] where Timestart >'"+timestart+"' and Timestart <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df7.to_excel(writer, sheet_name='ThreadVerification',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=ThreadVerification_Data.xls"})


@app.route('/download/MeasureDiameter',methods=['POST'])
def download_MeasureDiameter():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    print(timestart)
    df7 = pd.read_sql("SELECT*From[QC].[dbo].[MeasureDiameter] where Time_Finish >'"+timestart+"' and Time_Finish <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df7.to_excel(writer, sheet_name='MeasureDiameter',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=MeasureDiameter_Data.xls"})

@app.route('/download/Laser',methods=['POST'])
def download_laser():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT*From[QC].[dbo].[Laser] where TimeOutBarcode >'"+timestart+"' and TimeOutBarcode <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Laser_Data.xls"})
@app.route('/download/CNC',methods=['POST'])
def download_CNC():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT*From[QC].[dbo].[CNC] where Timeoutcnc2 >= '"+timestart+"' and Timeoutcnc2 <= '"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=CNC_Data.xls"})
@app.route('/download/ScanQR',methods=['POST'])
def download_ScanQR():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df7 = pd.read_sql("SELECT*From[QC].[dbo].[ScanQR] where Time_scan_tray >'"+timestart+"' and Time_scan_tray <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df7.to_excel(writer, sheet_name='ScanQR',index=False)
    # add headers
    writer.save()
    output.seek(0)
    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=ScanQR_Data.xls"})

@app.route('/download/TiltMeasurement',methods=['POST'])
def download_TiltMeasurement():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT Top(15000)* FROM [QC].[dbo].[TiltMeasurement] where Time_finish >'"+timestart+"' and Time_finish <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=TiltMeasurement_Data.xls"})
@app.route('/download/Airtight',methods=['POST'])
def download_airtight():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT * FROM  QC.dbo.air_tight  WHERE Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Airtight_Data.xls"})
@app.route('/download/AirtightChamfer',methods=['POST'])
def download_airtightChamfer():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT * FROM  QC.dbo.air_tight_chamfer  WHERE Time_Start >'"+timestart+"' and Time_Start <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)
 
    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Airtight_Chamfer_Data.xls"})
@app.route('/download/Airtightwindow',methods=['POST'])
def download_airtightwindow():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT * FROM  QC.dbo.air_tight_window  WHERE Time_Finish_2 >='"+timestart+"' and Time_Finish_2 <='"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)
 
    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Airtight_Window_Data.xls"})
@app.route('/download/AirguageQC',methods=['POST'])
def download_AirGuageQC():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT * FROM [QC].[dbo].[Measure_Diameter] where Time_finish >'"+timestart+"' and Time_finish <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=AirGuageQC_Data.xls"})
@app.route('/download/AirguageGC',methods=['POST'])
def download_AirGuageGC():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=GC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT * FROM [GC].[dbo].[Measure_Diameter] where Time_finish >'"+timestart+"' and Time_finish <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=AirGuageGC_Data.xls"})
@app.route('/download/Classification',methods=['POST'])
def download_classification():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT [Pallet].[SPLR_LOT_NO], [Classification].* From [Classification] left join [Pallet] on [Classification].pallet_num= [Pallet].Pallet_Name where [Classification].Time_check >='"+timestart+"' and [Classification].Time_check <='"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.to_excel(writer, sheet_name='layer1',index=False)

    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Classification_Data.xls"})
@app.route('/count',methods=['POST'])
def countsss():
    may = request.form['may']
    kq = request.form['kq']
    today = datetime.datetime.now()
    nextday = today + datetime.timedelta(days=1)
    timenw = datetime.datetime.now().strftime("%H:%M:%S")
    strtoday = today.strftime("'%Y-%m-%d'")
    strnextday = nextday.strftime("'%Y-%m-%d'")
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    df1 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = '"+may+"' and Result ='"+kq+"' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df2 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 2' and Result ='OK' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df3 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 3' and Result ='OK' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df4 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 4' and Result ='OK' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df5 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 5' and Result ='OK' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df6 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 1' and Result ='NG' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df7 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 2' and Result ='NG' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df8 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 3' and Result ='NG' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df9 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 4' and Result ='NG' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    # df10 = pd.read_sql("SELECT COUNT(*) From[QC].[dbo].[Laser] where MachineNo = 'MÁY 5' and Result ='NG' and TimeOutBarcode > "+strtoday+" and TimeOutBarcode < "+strnextday, conn)
    rsp = int(df1.loc[0])
    return str(rsp)

@app.route('/downloadtemp/Coating',methods=['POST'])
def download_temp_coating():
    global coatname
    try:
        # coatname = request.form['coatingpos']
        # print('b')
        # print(coatname)
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=Auto; Trusted_Connection=No;',timeout=1)
        timestart=request.form['timestart']
        timeend=request.form['timeend']
        coatname = request.form['coatingpos']
        print('a')
        print(coatname)
        #print(timestart)
        df2 = pd.read_sql("select * from data_"+coatname+"_W5 where Datetime >'"+timestart+"' and Datetime <'"+timeend+"'", conn)
        df1 = pd.read_sql("select Datetime, average_temp1, average_humid1 from data_"+coatname+"_W5 where Datetime >'"+timestart+"' and Datetime <'"+timeend+"' order by Datetime desc",conn)
        # print('abc:',"select Datetime, average_temp1, average_humid1 from data_"+coatname+"_W5 where Datetime >'"+timestart+"' and Datetime <'"+timeend+"' order by Datetime desc")
        df1["Datetime"] = pd.to_datetime(df1["Datetime"], format='%Y-%m-%d %H:%M:%S.%f')
        df1 = df1.set_index('Datetime')
        target = pd.read_sql("select * from [Auto].[dbo].[Settinglimitcoatingx5] where Coating = '"+coatname+"'",conn)
        temp_min = target.loc[0,'Tempmin']  # Thay thế bằng giá trị nhiệt độ mục tiêu của bạn
        temp_max = target.loc[0,'Tempmax']  # Thay thế bằng giá trị độ ẩm mục tiêu của bạn
        humid_min= target.loc[0,'Humidmin']
        humid_max= target.loc[0,'Humidmax']
        plt.figure(figsize=(10, 6))
        plt.scatter(df1.index, df1['average_temp1'], label='Average Temperature',color='orange',  s=5)
        plt.scatter(df1.index, df1['average_humid1'], label='Average Humidity',color='blue',  s=5)

        plt.axhline(y=temp_min, color='red', linestyle='-', label='Temp Min')
        plt.axhline(y=temp_max, color='green', linestyle='-', label='Temp Max')
        plt.axhline(y=humid_min, color='yellow', linestyle='-', label='Humidity Min')
        plt.axhline(y=humid_max, color='purple', linestyle='-', label='Humidity Max')

        plt.text(df1.index[0], temp_min, f'          {int(temp_min)}', color='r', va='bottom', ha='left')
        plt.text(df1.index[0], temp_max, f'          {int(temp_max)}', color='g', va='bottom', ha='left')
        plt.text(df1.index[0], humid_min, f'          {int(humid_min)}', color='yellow', va='bottom', ha='left')
        plt.text(df1.index[0], humid_max, f'          {int(humid_max)}', color='purple', va='bottom', ha='left')


        plt.xlabel('Datetime')
        plt.ylabel('Values')
        plt.title(coatname+' Average Temperature & Humidity',loc='left')
        plt.legend(bbox_to_anchor= (1, 1.1), ncol = 3,fontsize= 6)
        plt.xticks(rotation=45, fontsize= 8)
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d %H:%M'))  
        plt.tight_layout()
        plt.savefig('static/images/'+coatname+'.png')
        image_path = 'static/images/'+coatname+'.png'

        output = io.BytesIO()
        writer = pd.ExcelWriter(output,engine='xlsxwriter')    # add a sheet
        df2.index+=1
        df2.to_excel(writer, sheet_name=str(coatname))
        worksheet2 = writer.book.add_worksheet("Chart")

        # Get the xlsxwriter workbook and worksheet objects for the new sheet.
        workbook = writer.book
        worksheet2.insert_image('A1', image_path, {'x_offset': 15, 'y_offset': 10})

        # add headers
        writer.save()
        output.seek(0)
        

        return Response(output, mimetype="application/ms-excel",
                        headers={"Content-Disposition": "attachment;filename=Temp_Humid_Coating.xls"})
    except Exception as e:
        return e

@app.route('/download/Coating',methods=['POST'])
def download_coating():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=Auto; Trusted_Connection=No;',timeout=1)
    timestart=request.form['timestart']
    timeend=request.form['timeend']
    df1 = pd.read_sql("SELECT * FROM [Auto].[dbo].[Product_Quantity_W5_Coating1] where DateTime >'"+timestart+"' and DateTime <'"+timeend+"'", conn)
    df2 = pd.read_sql("SELECT * FROM [Auto].[dbo].[Product_Quantity_W5_Coating2] where DateTime >'"+timestart+"' and DateTime <'"+timeend+"'", conn)
    df3 = pd.read_sql("SELECT * FROM [Auto].[dbo].[Product_Quantity_W5_Coating3] where DateTime >'"+timestart+"' and DateTime <'"+timeend+"'", conn)
    df4 = pd.read_sql("SELECT * FROM [Auto].[dbo].[Product_Quantity_W5_Coating4] where DateTime >'"+timestart+"' and DateTime <'"+timeend+"'", conn)
    df5 = pd.read_sql("SELECT * FROM [Auto].[dbo].[Product_Quantity_W5_Coating5] where DateTime >'"+timestart+"' and DateTime <'"+timeend+"'", conn)
    df6 = pd.read_sql("SELECT * FROM [Auto].[dbo].[Product_Quantity_W5_Coating6] where DateTime >'"+timestart+"' and DateTime <'"+timeend+"'", conn)
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine='xlsxwriter')    # add a sheet
    df1.index+=1
    df2.index+=1
    df3.index+=1
    df4.index+=1
    df5.index+=1
    df6.index+=1
    df1.to_excel(writer, sheet_name='layer1')
    df2.to_excel(writer, sheet_name='layer2')
    df3.to_excel(writer, sheet_name='layer3')
    df4.to_excel(writer, sheet_name='layer4')
    df5.to_excel(writer, sheet_name='layer5')
    df6.to_excel(writer, sheet_name='layer6')
    # add headers
    writer.save()
    output.seek(0)

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=Coating_Data.xls"})
@app.route("/Electrical", methods=['POST', 'GET'])
def open_Electrical():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(2): 
            process="Electrical"
            conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
            
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM [QC].[dbo].[Electrical]")
            data =cursor.fetchall()
            print(data)
            return render_template('/pages/Electrical.html',user=user,process=process,eledata = data)
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Electrical/Chart", methods=['GET'])
def open_Electrical_chart():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(2): 
            process="Electrical"
            # timestart = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M")
            # timeend = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            # conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=E; Trusted_Connection=No;',timeout=1)
                
            # cursor = conn.cursor()
            # cursor.execute("SELECT top(24) * FROM [QC].[dbo].[Electrical]")
            # data =cursor.fetchall()
            return render_template('/pages/ElectricalChart.html',user=user,process=process)
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Electrical/<area>", methods=['GET'])
def open_Electrical_area(area):
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Electrical"
        if check_secr(2): 
            process="Electrical"
            return render_template('/pages/ElectricalArea.html',user=user,process=process,area=area)
        else:
            return render_template('pages/error.html',user=user)
        
    else:
        return redirect(url_for('login'))
@app.route("/Electrical/<area>/get", methods=['POST'])
def chart_Electrical(area):
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")

        # Lấy các thông số quan trọng
        process="Electrical"
        timestart = request.get_json()["timestart"]
        timeend = request.get_json()["timeend"]

        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM [Electric].[dbo].[ELec_main] where Name_Area = '"+area+"' and Time_get > '"+timestart+"' and Time_get < '"+timeend+"' order by Time_get" )
        data =cursor.fetchall() # là dữ liệu của bảng ELec_main
        cost = pd.read_sql("select * from [Electric].[dbo].[Elec_Cost] order by etype",conn)
        value = []
        time = []
        totalkw = 0
        money = []
        totalmn = 0
        if (data[-1][2]-data[0][2]).days >= 30:
            ctime = 0
            svalue = 0
            smoney = 0
            ctime='....'
            for i in range(len(data)):
                if data[i][2].day <=10:
                    ky = 'k1'
                elif data[i][2].day >20:
                    ky = 'k3'
                else:
                    ky = 'k2'                
                if ky != ctime[-2:]:
                    value.append(round(svalue,2))
                    money.append(round(smoney))
                    time.append(ctime)
                    totalkw+=svalue
                    totalmn+=smoney
                    ctime = data[i][2].strftime("%Y-%m-")+ky
                    svalue = data[i][15]
                    smoney = data[i][15] * cost.loc[data[i][17]-1,'cost']
                else:
                    svalue+=data[i][15]
                    smoney+= data[i][15] * cost.loc[data[i][17]-1,'cost']
                # total+=data[i][15]
            value.append(round(svalue,2))
            value = value[1:]
            money.append(round(smoney))
            money = money[1:]
            time.append(ctime)
            time = time[1:]
            totalkw+=svalue 
            totalmn+=smoney
        elif len(data) > 48:
            ctime = 0
            svalue = 0
            smoney = 0
            for i in range(len(data)):
                if data[i][2].strftime("%Y-%m-%d") != ctime:
                    value.append(round(svalue,2))
                    money.append(round(smoney))
                    time.append(ctime)
                    totalkw+=svalue
                    totalmn+=smoney
                    ctime = data[i][2].strftime("%Y-%m-%d")
                    svalue = data[i][15]
                    smoney = data[i][15] * cost.loc[data[i][17]-1,'cost']
                else:
                    svalue+=data[i][15]
                    smoney+= data[i][15] * cost.loc[data[i][17]-1,'cost']
            value.append(round(svalue,2))
            value = value[1:]
            money.append(round(smoney))
            money = money[1:]
            time.append(ctime)
            time = time[1:]
            totalkw+=svalue
            totalmn+=smoney
        else:
            for i in range(len(data)):
                value.append(round(data[i][15],2))
                money.append(round(data[i][15] * cost.loc[data[i][17]-1,'cost']))
                time.append(data[i][2].strftime("%Hh%Mm"))
                totalkw+=data[i][15]
                totalmn+=data[i][15] * cost.loc[data[i][17]-1,'cost']
        data = {}
        data["time"]=time
        data["kwh"] = {
            "name":area,
            "data":value,
            "color":'#1A56DB'
        }
        data["money"] = {
            "name":area,
            "data":money,
            "color":'#FDBA8C'
        }
        data["totalkw"] = "{:,}".format(round(totalkw,2))
        data["totalmn"] = "{:,}".format(round(totalmn))
        print(data["money"])
        return jsonify(data)
    else:
        return redirect(url_for('login'))
    
@app.route("/Electrical/Chart/get", methods=['POST'])
def chartElectrical():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="Electrical"
        timestart = request.get_json()["timestart"]
        timeend = request.get_json()["timeend"]
        # timestart = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M")
        # timeend = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
            
        cursor = conn.cursor()
        cursor.execute("SELECT  * FROM [Electric].[dbo].[ELec_main] where Time_get > '"+timestart+"' and Time_get < '"+timeend+"' order by Time_get" )
        data =cursor.fetchall()
        cost = pd.read_sql("select * from [Electric].[dbo].[Elec_Cost] order by etype",conn)
        areaname = pd.read_sql("select distinct(Name_Area) from [Electric].[dbo].[ELec_main] where Name_Area not like '%All%' and Time_get > '"+timestart+"' and Time_get < '"+timeend+"'",conn)
        color = ["#FF0000", "#00FFFF", "#0000FF", "#00008B", "#ADD8E6", "#800080", "#FFFF00", "#00FF00", "#FFC0CB", "#C0C0C0", "#808080", "#000000", "#FFA500", "#A52A2A", "#800000", "#008000", "#808000", "#7FFFD4"]
        data = {}
        data["labels"] = []
        data["colors"] =[]
        data["kwh"] = []        
        data["money"] = []
        allkw = 0
        allmn = 0
        for i in range(len(areaname)):
            totalkw = 0
            totalmn = 0
            area = areaname.loc[i,"Name_Area"]
            cursor.execute("SELECT  Ea_Sub,etype FROM [Electric].[dbo].[ELec_main]  where Name_Area = '"+area+"' and Time_get > '"+timestart+"' and Time_get < '"+timeend+"' order by Time_get" )
            dt = cursor.fetchall()
            for j in range(len(dt)):
                totalkw+=dt[j][0]
                totalmn+=dt[j][0] * cost.loc[dt[j][1]-1,'cost']
            allkw+=totalkw
            allmn+=totalmn
            data["kwh"].append(round(totalkw,2))
            data["money"].append(round(totalmn))
            data["labels"].append(area)
            data["colors"].append(color[i])
        data["allkw"] = "{:,}".format(round(allkw,2))
        data["allmn"] = "{:,}".format(round(allmn))
        print(data["money"])
        return jsonify(data)
    else:
        return redirect(url_for('login'))
    
@app.route("/Electrical/update", methods=['POST', 'GET'])
def update_Electrical():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    
    data = pd.read_sql("SELECT * FROM [QC].[dbo].[Electrical]",conn)
    edata = {}
    for i in range(len(data)):
        edata[data.loc[i,"Area"]]={
            "Uavg":data.loc[i,"Voltage_avg"],
            "Uab":data.loc[i,"Uab"],
            "Ubc":data.loc[i,"Ubc"],
            "Uca":data.loc[i,"Uca"],
            "Total":data.loc[i,"Total"],
        }
    # data = data.to_json()
    return jsonify(edata)
@app.route("/Electrical/Cost", methods=['GET'])
def open_Electrical_cost():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(3): 
            process="Electrical"
            # timestart = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M")
            # timeend = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM [Electric].[dbo].[Elec_Cost] order by etype")
            data =cursor.fetchall()

            return render_template('/pages/ElectricalCost.html',user=user,process=process,ohour=data[0][2],nhour=data[1][2],rhour=data[2][2])
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Electrical/SaveCost", methods=['POST'])
def open_Electrical_scost():
    if request.cookies.get("login_status")=='ok':
        try:
            user = request.cookies.get("user_name")
            process="Electrical"
            ohour = request.get_json()["ohour"]
            nhour = request.get_json()["nhour"]
            rhour = request.get_json()["rhour"]
            # timestart = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M")
            # timeend = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
                
            cursor = conn.cursor()
            cursor.execute("update [Electric].[dbo].[Elec_Cost] set cost = "+str(ohour)+"where etype=1")
            cursor.commit()
            cursor.execute("update [Electric].[dbo].[Elec_Cost] set cost = "+str(nhour)+"where etype=2")
            cursor.commit()
            cursor.execute("update [Electric].[dbo].[Elec_Cost] set cost = "+str(rhour)+"where etype=3")
            cursor.commit()

            return "OK"
        except:
            return "Fail"
    else:
        return redirect(url_for('login'))
@app.route("/HeatTreatment", methods=['POST', 'GET'])
def open_heat():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="HeatTreatment"
        if request.method == 'POST':
            timestart=request.form['daystart']+" "+request.form['timestart']
            timeend=request.form['dayend']+" "+request.form['timeend']
            ma = request.form['code']
            conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=SPC; Trusted_Connection=No;',timeout=1)
            cursor = conn.cursor()
            # data = laser_all_data("","","",timestart,timeend)
            if len(ma)==8:
                
                try:
                    ccode = ma[-3:]
                    ctime = ma[:-4]
                    if ctime[1].upper()=='X':
                        cmonth = 10
                    elif ctime[1].upper()=='Y':
                        cmonth = 11
                    elif ctime[1].upper()=='Z':
                        cmonth = 12
                    else:
                        cmonth = '0'+ctime[1]
                    rtime = f'202{ctime[0]}-{cmonth}-{ctime[-2:]}'
                    print(rtime)
                    print("SELECT  *  FROM [SPC].[dbo].[HXL_HeatTreatment_Furnance] where Furnance_Code = '"+ccode+"' and TimeSet = '"+rtime+"'")
                    cursor.execute("SELECT  *  FROM [SPC].[dbo].[HXL_HeatTreatment_Furnance] where Furnance_Code = '"+ccode+"' and TimeSet = '"+rtime+"'")
                    
                    data = cursor.fetchone()
                    hcode = data[1]
                except:
                    hcode = 'nocode'
                cursor.execute("SELECT Top(15000)* From [SPC].[dbo].[HXL_HeatTreatment_FileCode] where HeatTreatment_Code like '%"+hcode+"%'")
            else:
                cursor.execute("SELECT Top(15000)* From [SPC].[dbo].[HXL_HeatTreatment_FileCode] where Time_Out >'"+timestart+"' and Time_In <'"+timeend+"'")
            data =cursor.fetchall()
            # return render_template('/pages/Heat2.html',user=user,process=process,timestart="",timeend="",code=ma)
            return render_template('pages/Heat2.html',user=user,process=process,post_flag = True,data = data, timestart=timestart,timeend=timeend,code=ma)
        else:
            return render_template('/pages/Heat2.html',user=user,process=process,timestart="",timeend="",code="")
    else:
        return redirect(url_for('login'))
@app.route("/HeatTreatment/realtime", methods=['POST', 'GET'])
def open_heatrt():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="HeatTreatment"
        return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/Heat2/<filename>", methods=['POST', 'GET'])
def heat2(filename):
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=LWM; Trusted_Connection=No;',timeout=1)
    cursor = conn.cursor()
    cursor.execute("SELECT Time_In,Time_Indicated,Time_Out,Machine FROM [SPC].[dbo].[HXL_HeatTreatment_FileCode] where HeatTreatment_Code like '%"+filename+"%'")
    data1 = cursor.fetchone()
    # data = pd.read_csv('static/csv/'+filename,skiprows=8,sep=';')
    print("SELECT  * FROM [SPC].[dbo].[HXL_HeatTreatment_Data] where  DateTimeData >'"+str(data1[0])+"' and DateTimeData < '"+str(data1[2])+"'  and Machine ='"+data1[3]+"' order by id desc")
    cursor.execute("SELECT  * FROM [SPC].[dbo].[HXL_HeatTreatment_Data] where DateTimeData >'"+str(data1[0])+"' and DateTimeData < '"+str(data1[2])+"' and Machine ='"+data1[3]+"' order by id desc" )
    
    data =cursor.fetchall()
    value = []
    value2 = []
    value3 = []
    time = []
    money = []
    for i in range(len(data)):
        value.append(round(data[len(data)-i-1][3]))
        value2.append(round(data[len(data)-i-1][4]))
        value3.append(round(data[len(data)-i-1][5]))
        money.append(round(data[len(data)-i-1][5]))
        time.append(int(data[len(data)-i-1][2].timestamp()*1000))
    data = {}
    # time_from_sql = [datetime.datetime.strptime(str(t), '%Y-%m-%d %H:%M:%S') for t in time]

    # Tạo danh sách thời gian có định dạng chuẩn để truyền vào JavaScript
    # time_for_js = [datetime.datetime.strptime(t[:-7], '%Y-%m-%d %H:%M:%S').isoformat() for t in time]
    data["time"]=time
    data["kwh"] = [{
        "name":"MainTemp",
        "data":value,
        "color":'#1A56DB'
    },
    {
        "name":"OverTemp",
        "data":value2,
        "color":'#FDBA8C'
    },
    {
        "name":"Cooling",
        "data":value3,
        "color":'#250998'
    }]
    data["money"] = {
        "name":filename,
        "data":money,
        "color":'#FDBA8C'
    }
    data["totalkw"] = "0"
    data["totalmn"] = "0"
    data["Indicated"] =  int(data1[1].timestamp()*1000)
    print(data["Indicated"])
    print(len(data["kwh"][0]["data"]))
    return jsonify(data)
@app.route("/CMM", methods=['POST', 'GET'])
def open_CMM():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="CMM"
        if check_secr(5): 
            if request.method == 'POST':
                timestart=request.form['daystart']+" "+request.form['timestart']
                timeend=request.form['dayend']+" "+request.form['timeend']
                Product = request.form['mahang']
                Machine = request.form['machine']
                dmc = request.form['dmc']
                conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
                cursor = conn.cursor()
                cursor.execute("SELECT TimeStart,TimeSave,requester,  CodePurpose, DMC, Line, CMMmachine,CMMCode, Operator, (CASE WHEN (SELECT count(value) FROM STRING_SPLIT(Product, ' ')) >= 2 THEN (SELECT dbo.ConcatValuesInColumn(Product, 2)) ELSE Product END) as Product, SUM(CASE WHEN Result != '' THEN 1 ELSE 0 END) AS Total,SUM(CASE WHEN id like '%No Tol%' THEN 1 ELSE 0 END) AS NoTol, SUM(CASE WHEN id like '%In Tol%' THEN 1 ELSE 0 END) AS InTol,SUM(CASE WHEN Result = 'OK'  and id not like '%No Tol%' and id not like '%In Tol%' THEN 1 ELSE 0 END) AS OK_Total,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' THEN 1 ELSE 0 END) AS NG_Total,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and (id like '% C\_%' ESCAPE '\\' or id like '% C.%' or id like '% C %' or id like '% CC %') THEN 1 ELSE 0 END) AS OK_C,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and (id like '% C\_%' ESCAPE '\\' or id like '% C.%'or id like '% C %' or id like '% CC %') THEN 1 ELSE 0 END) AS NG_C,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and id like '%CM%' THEN 1 ELSE 0 END) AS OK_CM,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and id like '%CM%' THEN 1 ELSE 0 END) AS NG_CM,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and id not like '%CM%' and id not like '% C\_%' ESCAPE '\\' and id not like '% C.%' and id not like '% C %' and id not like '% CC %' THEN 1 ELSE 0 END) AS OK_MM,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and id not like '%CM%' and id not like '% C\_%' ESCAPE '\\' and id not like '% C.%' and id not like '% C %' and id not like '% CC %' THEN 1 ELSE 0 END) AS NG_MM FROM CMMdata where timesave >'"+timestart+"' and timesave <'"+timeend+"' and dmc like '%"+dmc+"%' and product like '%"+Product+"%' and CMMmachine like '%"+Machine+"%' GROUP BY DMC,TimeStart,requester, CodePurpose,CMMCode, Line, CMMmachine, Operator, TimeSave, Product")
                data1 =cursor.fetchall()
                soluong = len(data1)
                print(data1)
                return render_template('/pages/CMMdata.html',post_flag = True,soluong=soluong,process=process,user=user,data = data1,timestart=timestart,timeend=timeend,mahang = Product, machine=Machine,dmc=dmc)
            else:
                return render_template('/pages/CMMdata.html',user=user,process=process,timestart="",timeend="",dmc="")   
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/CMMpurpose_update", methods=['POST'])
def CMM_update():
    data = request.get_json()
    print(data)
    dmc = data.get("dmc")
    timesave = data.get("timesave")
    purpose = data.get("purpose")
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    if check_secr(6): 
        cursor = conn.cursor()
        print("Update [QC].[dbo].[CMMdata] set CodePurpose ='"+purpose+"' where DMC ='"+dmc+"' and timesave = '"+timesave[:-3]+"'")
        cursor.execute("Update [QC].[dbo].[CMMdata] set CodePurpose ='"+purpose+"' where DMC ='"+dmc+"' and timesave = '"+timesave[:-3]+"'")
        cursor.commit()
        return "OK"
    else:
        return  "You don't have permission"
@app.route("/CMM_Delete", methods=['POST'])
def CMM_Delete():
    data = request.get_json()
    dmc = data.get("dmc")
    timesave = data.get("timess")
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
    if check_secr(7): 
        cursor = conn.cursor()
        # print("Delete from [QC].[dbo].[CMMdata]  where DMC ='"+dmc+"' and timesave = '"+timesave[:-3]+"'")
        cursor.execute("Delete from [QC].[dbo].[CMMdata] where DMC ='"+dmc+"' and timesave = '"+timesave[:-3]+"'")
        cursor.commit()
        return "OK"
    else:
        return  "You don't have permission"
@app.route("/CMM/Chart", methods=['GET'])
def open_CMM_chart():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(10): 
            process="CMM"
            # timestart = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M")
            # timeend = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            # conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=E; Trusted_Connection=No;',timeout=1)
                
            # cursor = conn.cursor()
            # cursor.execute("SELECT top(24) * FROM [QC].[dbo].[Electrical]")
            # data =cursor.fetchall()
            return render_template('/pages/cmmchart2.html',user=user,process=process)
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/CMM/Chart/get", methods=['POST', 'GET'])
def open_CMMchart():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        timestart = request.get_json()["timestart"]
        timeend = request.get_json()["timeend"]
        cmmmachine = str(request.get_json()["cmmmachine"]).replace('[','(').replace(']',')')
        # timestart = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M")
        # timeend = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        print("select timestart, timesave, cmmmachine from [QC].[dbo].[CMMdata] where cmmmachine in "+cmmmachine+" and timesave > '"+timestart+"' and timestart < '"+timeend+"' group by  timestart, timesave, cmmmachine order by cmmmachine,timesave")
        
        cursor = conn.cursor()
        cursor.execute("select timestart, timesave, cmmmachine from [QC].[dbo].[CMMdata] where cmmmachine in "+cmmmachine+" and timesave > '"+timestart+"' and timestart < '"+timeend+"' group by  timestart, timesave, cmmmachine order by cmmmachine,timesave" )
        data1 =cursor.fetchall()
        data = {}
        data["data"] = []
        for dt in data1:
            data["data"].append(
                {
                    "x": dt[2],
                    "y": [int(dt[0].timestamp()*1000+25200000), int(dt[1].timestamp()*1000+25200000)]
                }
            )
        print(data)
        return jsonify(data)
    else:
        return redirect(url_for('login'))
@app.route("/CMM/Form", methods=['POST', 'GET'])
def open_CMMform():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        if check_secr(8): 
            process="CMM"
            conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1) 
            cursor = conn.cursor()
            cursor.execute("select distinct(FormName) from CMMFormdata")
            data1 =cursor.fetchall()
            if request.method == 'POST':
                return render_template('/pages/CMMform.html',post_flag = True,process=process,user=user)
            else:
                return render_template('/pages/CMMform.html',user=user,process=process,formname = data1)   
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/CMM/Form/Manager", methods=['GET','POST'])
def open_CMMformmanager():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="CMM"        
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
        cursor = conn.cursor()
        if check_secr(9): 
            if request.method == 'POST':
                oldform=request.form['oldform']
                formname=request.form['formName']
                filename = request.form['filename']
                os.makedirs('static/cmmform/'+formname)
                shutil.copy('static/cmmform/'+oldform+'/'+filename, 'static/cmmform/'+formname+'/'+filename)
                if len(pd.read_sql("select id from CMMFormdata where FormName = '"+formname+"'",conn)):
                    cursor.execute("select distinct(FormName), [FileName] from CMMFormdata group by FormName,[FileName]")
                    data1 =cursor.fetchall()
                    return render_template('/pages/CMMformmanager.html',post_flag = True,process=process,user=user,data=data1) 
                cursor.execute("INSERT INTO [CMMFormData] (FormName, FileName, CircleNum, Characteristic, Link, Position) SELECT '"+formname+"' AS FormName,FileName,CircleNum,Characteristic,Link,Position FROM [CMMFormData] WHERE FormName = '"+oldform+"'")
                cursor.commit()
            cursor.execute("select distinct(FormName), [FileName] from CMMFormdata group by FormName,[FileName]")
            data1 =cursor.fetchall()
            return render_template('/pages/CMMformmanager.html',post_flag = True,process=process,user=user,data=data1) 
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/CMM/Form/Manager/Delete", methods=['POST'])
def open_CMMformdelete():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="CMM"
        dmc = request.get_json()["dmc"]
        password = request.get_json()["password"]
        shutil.rmtree('static/cmmform/'+dmc)
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)   
        confirm_user = pd.read_sql("SELECT Top(1) * FROM [Auto].[dbo].[Web_User] where username = '"+user+"' and password = '"+password+"'" , conn)
        if len(confirm_user) == 1:
            cursor = conn.cursor()
            cursor.execute("delete from cmmformdata where formname = '"+dmc+"'")
            cursor.commit()
            return "OK"
        else:
            return "Wrong Password"     
    else:
        return redirect(url_for('login'))
@app.route("/CMM/Form/Manager/<formname>", methods=['POST', 'GET'])
def popupCMM1(formname):
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("SELECT * FROM [QC].[dbo].[CMMFormdata] where FormName = '"+formname+"'",conn)
    edata = {}
    for i in range(len(data)):
        edata[str(data.loc[i,"Position"])]={
            "Position":str(data.loc[i,"Position"]),
            "CircleNum":str(data.loc[i,"CircleNum"]),
            "Characteristic":data.loc[i,"Characteristic"],
            "Link":data.loc[i,"Link"]
        }
    # data = data.to_json()
    return jsonify(edata)
@app.route('/CMM/upload/<ref>', methods=['POST'])
def upload_file(ref):
    if 'file' not in request.files:
        print('No file part')
        return 'No file part'
    
    file = request.files['file']
    if file.filename == '':
        print('No selected file')
        return 'No selected file'

    # Xử lý file Excel tại đây (ví dụ: lưu file vào thư mục, đọc dữ liệu từ file, ...)
    # Dưới đây là ví dụ lưu file vào thư mục uploads
    file_path = 'static/cmmform/' + file.filename
    file.save(file_path)
    print('File uploaded successfully')
    
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Xử lý dữ liệu và trả về kết quả
    edata = {}
    row_number = 24  # Hàng bắt đầu từ 23 (C23, D23)
    ref_flag = False
    if ref != "No Reference":
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)  
        olddata = pd.read_sql("select * from cmmformdata where formname = '"+ref+"'",conn)
        ref_flag = True
    while True:
        num = sheet.cell(row=row_number, column=4).value
        charr = sheet.cell(row=row_number, column=5).value
        norminal = sheet.cell(row=row_number, column=7).value
        lower = sheet.cell(row=row_number, column=8).value
        upper = sheet.cell(row=row_number, column=9).value
        if num is None:
            break
        link = ''
        if ref_flag:
            for i in range(len(olddata)):
                if str(olddata.loc[i,"CircleNum"]) == str(num):
                    link = str(olddata.loc[i,"Link"])
                    break
        edata[row_number-23] = {"Num":num,
                                "Char":charr,
                                "Norminal":norminal,
                                "Lower":lower,
                                "Upper":upper,
                                "Link":link
                                }
        row_number += 1

    # Xóa file Excel sau khi hoàn thành
    # os.remove(file_path)
    print(edata)
    return jsonify(edata)
@app.route('/CMM/uploadtxt', methods=['POST'])
def upload_txtfile():
    dmc = request.get_json()["dmc"]
    print(dmc)
    if dmc =="":
        dmc= "datnon"
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("SELECT id,actual FROM [QC].[dbo].[CMMdata] where DMC like '%"+dmc+"%' order by IDx",conn)
    print("SELECT id,actual FROM [QC].[dbo].[CMMdata] where DMC = '"+dmc+"' order by IDx")
    edata = {}
    for i in range(len(data)):
        edata[i]={
            "id":str(data.loc[i,"id"]),
            "Actual":data.loc[i,"actual"]
        }
    # data = data.to_json()
    return jsonify(edata)
# @app.route('/CMM/uploadtxt', methods=['POST'])
# def upload_txtfile():
#     if 'file' not in request.files:
#         print('No file part')
#         return 'No file part'
    
#     file = request.files['file']
#     if file.filename == '':
#         print('No selected file')
#         return 'No selected file'

#     # Xử lý file Excel tại đây (ví dụ: lưu file vào thư mục, đọc dữ liệu từ file, ...)
#     # Dưới đây là ví dụ lưu file vào thư mục uploads
#     file_path = 'static/cmmform/' + file.filename
#     file.save(file_path)
#     print('File uploaded successfully')
    
#     with open(file_path, 'r') as file:
#         lines = file.readlines()
#     # Tách các giá trị trong mỗi dòng bằng ký tự tab và tạo danh sách các cột
#     data = [line.strip().split('\t') for line in lines]
#     data = configdata(data)
#     # Tạo DataFrame từ danh sách các dòng và đặt tên cột
#     df = pd.DataFrame(data[1:], columns=data[0])
#     edata= {}
#     for i in range(len(df)):
#         edata[i] = {"id":df.loc[i,"id"],
#                     "Actual":df.loc[i,"actual"]
#                     }

#     # Xóa file Excel sau khi hoàn thành
#     # os.remove(file_path)
#     print(edata)
#     return jsonify(edata)
@app.route("/CMM/<dmc>", methods=['POST', 'GET'])
def popupCMM(dmc):
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("SELECT * FROM [QC].[dbo].[CMMdata] where DMC = '"+dmc+"'",conn)
    edata = {}
    for i in range(len(data)):
        edata[data.loc[i,"id"]]={
            "Index":str(data.loc[i,"IDx"]),
            "TimeSave":str(data.loc[i,"TimeSave"])[:-7],
            "Actual":data.loc[i,"actual"],
            "Nominal":data.loc[i,"nominal"],
            "Upper":data.loc[i,"uppertol"],
            "Lower":data.loc[i,"lowertol"],
            "Unit":data.loc[i,"unit"],
            "Result":data.loc[i,"Result"]
        }
    # data = data.to_json()
    return jsonify(edata)
@app.route("/CMM/<dmc>/<code>", methods=['POST', 'GET'])
def popupCMM22(dmc,code):
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("SELECT * FROM [QC].[dbo].[CMMdata] where DMC = '"+dmc+"' and TimeSave = '"+code[:-3]+"'",conn)
    edata = {}
    for i in range(len(data)):
        edata[str(data.loc[i,"IDx"])]={
            "Index":str(data.loc[i,"IDx"]),
            "TimeSave":str(data.loc[i,"TimeSave"])[:-7],
            "Actual":data.loc[i,"actual"],
            "Nominal":data.loc[i,"nominal"],
            "Upper":data.loc[i,"uppertol"],
            "Lower":data.loc[i,"lowertol"],
            "Unit":data.loc[i,"unit"],
            "Result":data.loc[i,"Result"],
            "Name":data.loc[i,"id"]
        }
    # data = data.to_json()
    return jsonify(edata)
@app.route("/getForm", methods=['POST', 'GET'])
def getForm():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("select distinct(FormName) as Fname from CMMFormData",conn)
    edata = {}
    edata["FormName"]=[]
    for i in range(len(data)):
        edata["FormName"].append(data.loc[i,"Fname"])
    # data = data.to_json()
    return jsonify(edata)
@app.route("/getProduct", methods=['POST', 'GET'])
def getProduct():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("select distinct(Product) as Product from CMMData  order by Product",conn)
    edata = {}
    edata["Product"]=[]
    for i in range(len(data)):
        edata["Product"].append(data.loc[i,"Product"])
    # data = data.to_json()
    return jsonify(edata)
@app.route("/getMachine", methods=['POST', 'GET'])
def getMachine():
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("select distinct(CMMmachine) as Product from CMMData order by CMMmachine",conn)
    edata = {}
    edata["Product"]=[]
    for i in range(len(data)):
        edata["Product"].append(data.loc[i,"Product"])
    # data = data.to_json()
    return jsonify(edata)
@app.route("/CMM/export", methods=['POST', 'GET'])
def exportCMM():
    try:
        payload = request.get_json()
        formname = payload["fname"]
        dmclist1 = payload.get('dmc')   
        # dmclist = list(set(dmclist))
        dmclist = []
        for item in dmclist1:
            if item not in dmclist:
                dmclist.append(item)
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)   
        filename = pd.read_sql("select top 1 filename from cmmformdata where Formname ='"+formname+"'",conn).loc[0,"filename"]
        wb_sample = load_workbook('static/cmmform/'+formname+'/'+filename)
        sheet=wb_sample.active
        my_red = openpyxl.styles.colors.Color(rgb='00ff5555')
        my_gray = openpyxl.styles.colors.Color(rgb='00d9d9d9')
        daycl = openpyxl.styles.colors.Color(rgb='0055ff55')
        nightcl = openpyxl.styles.colors.Color(rgb='00FFA500')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        my_fill2 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_gray)
        dayfill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=daycl)
        nightfill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=nightcl)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border1 = Border(left=Side(style='dotted'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='dotted'))
        
        if len(dmclist) > 3:
            try:
                sheet.unmerge_cells(start_row=21,end_row=22,start_column=19,end_column=19)
            except:
                pass
            sheet.insert_cols(17,len(dmclist)-3)
            for i in range(len(dmclist)-3):
                
                for row_num in range(23, sheet.max_row + 1):
                    sheet.cell(row_num,17+i).border = border1
                    sheet.cell(row_num,17+i).alignment = Alignment(horizontal='center',vertical='center')
                    sheet.cell(row_num,17+i).font = Font(name="Arial")
                for row in range(2,10):
                    sheet.cell(row,17+i).border = Border(left=Side(style='thin'), right=Side(style='dashed'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in range(11,21):
                    sheet.cell(row,17+i).border = Border(top=Side(style='dashed'))
                sheet.cell(10,17+i).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
                sheet.cell(21,17+i).border = border
                sheet.cell(22,17+i).border = border
                sheet.cell(23,17+i).border = border
                sheet.cell(21,17+i).alignment = Alignment(horizontal='center',vertical='center')
                sheet.cell(22,17+i).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                sheet.cell(23,17+i).alignment = Alignment(horizontal='center',vertical='center')
                sheet.cell(21,17+i).font = Font(name="Arial")
                sheet.cell(22,17+i).font = Font(name="Arial")
                sheet.cell(23,17+i).font = Font(name="Arial")
                sheet.cell(22,17+i).fill = my_fill2
            # sheet.merge_cells(start_row=21,end_row=22,start_column=15+len(dmclist),end_column=15+len(dmclist))
    
        for idmc in range(len(dmclist)):
            try:
                # print("SELECT [DMC],[Line],[TimeSave],[id],[actual],[Result],[CodePurpose] FROM [QC].[dbo].[CMMdata] where dmc = '"+dmclist[idmc]+"'")
                idlist = pd.read_sql("select link as idx from cmmformdata where formname = '"+formname+"' order by Position",conn)
                data = pd.read_sql("SELECT [DMC],[Line],[TimeSave],[id],[actual],[Result],[CodePurpose] FROM [QC].[dbo].[CMMdata] where dmc = N'"+dmclist[idmc]+"'",conn)
                print(idmc,len(dmclist))
                if data.loc[0,"TimeSave"].hour < 8 or data.loc[0,"TimeSave"].hour > 19:
                    sheet.cell(21,idmc+16).fill = nightfill
                else:
                    sheet.cell(21,idmc+16).fill = dayfill
                sheet.cell(21,idmc+16).value = str(data.loc[0,"TimeSave"])[:-7]
                sheet.cell(22,idmc+16).value = data.loc[0,"Line"] +' '+ dmclist[idmc]
                print(data.loc[0,"Line"] +' '+ dmclist[idmc])
                sheet.cell(23,idmc+16).value = data.loc[0,"CodePurpose"]
                sheet.cell(23,idmc+16).border = border
                data.set_index("id", inplace=True)
                for i in range(len(idlist)):
                    try:
                        sheet.cell(24+i,idmc+16).value = round(float(data.loc[idlist.loc[i,"idx"],"actual"]),3)
                        sheet.cell(24+i,idmc+16).number_format = '0.000'
                        if "No Tol" not in idlist.loc[i,"idx"]:                            
                            if data.loc[idlist.loc[i,"idx"],"Result"].strip() == "":
                                if data.loc[idlist.loc[i,"idx"].replace(".XA",'').replace(".YA",'').replace(".ZA",'').replace(".R",'').replace(".PH",''),"Result"].strip() != "OK":
                                    sheet.cell(24+i,idmc+16).fill = my_fill
                            elif data.loc[idlist.loc[i,"idx"],"Result"].strip() != "OK":
                                sheet.cell(24+i,idmc+16).fill = my_fill
                    except Exception as e:
                        sheet.cell(24+i,idmc+16).fill = my_fill
                    
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"nominal"]
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"lowertol"]
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"uppertol"]
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"unit"]
                        # sheet.cell(3+i,idmc+15).value = "CMM Machine"
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"actual"]
                
            except:
                print(dmclist[idmc])
        # data = data.to_json()
        output = io.BytesIO()

        # add headers
        wb_sample.save(output)
        # writer.save()
        output.seek(0)

        return Response(output, mimetype="application/ms-excel",
                        headers={"Content-Disposition": "attachment;filename=CMM_Data.xls"})
    except:
        f = open("loi.txt",'a')
        f.write(traceback.format_exc())
        f.close()
@app.route("/CMM/createform", methods=['POST'])
def createform():
    payload = request.get_json()
    print(payload)
    FormName = payload["FormName"]
    filename = payload["filename"]
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
    data = pd.read_sql("SELECT count(FormName) as count From CMMFormData where FormName = '"+FormName+"'",conn)
    if data.loc[0,"count"] > 0:
        return "Form Name already exists"
    os.makedirs('static/cmmform/'+FormName)
    shutil.copy('static/cmmform/'+filename, 'static/cmmform/'+FormName+'/'+filename)
    cursor = conn.cursor()
    count = 0
    for info in payload["Data"][1:]:
        cursor.execute( 
            '''INSERT INTO CMMFormData VALUES (?,?,?,?,?,?) ''',
                FormName, 
                filename,
                info["CircleNum"], 
                info["Characteristic"],    
                info["Link"],
                count,
        )
        count +=1
    cursor.commit()
    return "OK"
@app.route("/CMM/editform", methods=['POST'])
def editform():
    payload = request.get_json()
    print(payload)
    FormName = payload["FormName"]
    filename = payload["filename"]
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1) 
    cursor = conn.cursor()
    count = 0
    cursor.execute("delete from cmmformdata where formname = '"+FormName+"'")
    cursor.commit()
    for info in payload["Data"][1:]:
        cursor.execute( 
            '''INSERT INTO CMMFormData VALUES (?,?,?,?,?,?) ''',
                FormName, 
                filename,
                info["CircleNum"], 
                info["Characteristic"],    
                info["Link"],
                count,
        )
        count +=1
    cursor.commit()
    return "OK"
@app.route("/User", methods=['GET','POST'])
def user():
    if request.cookies.get("login_status")=='ok':
        user = request.cookies.get("user_name")
        process="CMM"        
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
        cursor = conn.cursor()
        if check_secr(0):
            cursor.execute("select role from [Auto].[dbo].[Web_Role] order by role")
            data1 =cursor.fetchall()
            cursor.execute("select id,access,arrange FROM [Auto].[dbo].[Web_AccessList] order by arrange")
            data2 =cursor.fetchall()
            return render_template('/pages/user.html',process=process,user=user,data=data1,accesslist = data2) 
        else:
            return render_template('pages/error.html',user=user)
    else:
        return redirect(url_for('login'))
@app.route("/useraccess/<role>", methods=['POST', 'GET'])
def useraccess(role):
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)    
    data = pd.read_sql("select security FROM [Auto].[dbo].[Web_Role] where role = '"+role+"'",conn)
    # edata = {}
    # for i in range(len(data)):
    #     edata[data.loc[i,"access"]]= str(data.loc[i,"id"])
    # # data = data.to_json()
    return str(data.loc[0,"security"])
@app.route("/setsecurity", methods=['POST', 'GET'])
def setsecurity():
    payload = request.get_json()
    print(payload)
    if payload["security"] % 2 == 1:
        scr = "1"
    else:
        scr = str(payload["security"])
    conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)   
    cursor = conn.cursor()
    cursor.execute("update [Auto].[dbo].[Web_Role] set security ='"+scr+"' where role = '"+payload["role"]+"'")
    cursor.commit()
    # FormName = payload["FormName"]
    # filename = payload["filename"]
    # conn = pyodbc.connect('Dri/t security FROM [Auto].[dbo].[Web_Role] where role = '"+role+"'",conn)
    # edata = {}
    # for i in range(len(data)):
    #     edata[data.loc[i,"access"]]= str(data.loc[i,"id"])
    # # data = data.to_json()
    return "OK"
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0',port=8055)
